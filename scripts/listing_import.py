#!/usr/bin/env python3
"""남의 매물장 엑셀 → 콕집 매물장.

중개사무소마다 매물장 서식이 다르다. 열 이름도 순서도 단위도 제각각이고, 제목 줄이
위에 붙거나 동·호가 한 칸에 들어 있기도 하다. 그래서 '이 서식만 받는다'가 아니라
**아무 서식이나 받아 우리 칸으로 옮기는** 쪽으로 만든다.

옮기는 순서는 셋이다.
  ① 열 이름으로 맞춘다 — 중개사들이 실제로 쓰는 말을 모아 뒀다(단지·물건지·매가…).
  ② 이름으로 못 맞춘 열은 **값을 보고** 맞춘다 — 전화번호 꼴, 층 표기('3/15'),
     거래유형 낱말이 들어 있으면 그 열이 무엇인지 알 수 있다.
  ③ 단위는 열 단위로 정한다 — 같은 '85,000' 이 매매가면 8억 5천이고 월세면 8억 5천이
     아니다. 열의 값 분포와 머리글을 같이 보고 정하고, **무엇으로 읽었는지 밝힌다.**

사람이 확인하기 전에는 저장하지 않는다. analyze() 는 무엇을 어떻게 읽었는지(열 매핑·
단위·건너뛴 줄)를 전부 돌려주고, 저장은 확인을 마친 뒤 따로 한다.
"""
from __future__ import annotations

import csv
import io
import json
import re
from datetime import date, datetime, time as _dtime

# ── 우리 칸 ──────────────────────────────────────────────────────────────────
# (키, 화면 이름, 값 종류). 값 종류가 파싱 방법과 단위 추론을 가른다.
FIELDS: list[tuple[str, str, str]] = [
    ("trade_type", "거래유형", "trade"),
    ("type", "물건종류", "ptype"),
    ("complex_name", "단지·건물명", "text"),
    ("address", "주소", "text"),
    ("dong", "동", "unit"),
    ("ho", "호", "unit"),
    ("price", "매매가", "money"),
    ("deposit", "보증금", "money"),
    ("rent_price", "월세", "money"),
    ("premium", "권리금", "money"),
    ("maintenance_fee", "관리비", "money"),
    ("loan_amount", "융자", "money"),
    ("area2_m2", "전용면적", "area"),
    ("area1_m2", "공급면적", "area"),
    ("land_area_m2", "대지면적", "area"),
    ("total_area_m2", "연면적", "area"),
    ("floor_info", "층", "floor"),
    ("total_floor", "총층", "int"),
    ("direction", "향", "text"),
    ("room_cnt", "방", "int"),
    ("bath_cnt", "욕실", "int"),
    ("parking", "주차", "num"),
    ("move_in", "입주", "text"),
    ("settle_ymd", "잔금", "ymd"),
    ("approve_ymd", "준공", "ymd"),
    ("owner_name", "소유자", "text"),
    ("owner_tel", "소유자 연락처", "tel"),
    ("contact", "연락처", "tel"),
    ("manager", "담당자", "text"),
    ("current_biz", "업종", "text"),
    ("land_category", "지목", "text"),
    ("land_use", "용도지역", "text"),
    ("tenant_until", "임차 만기", "text"),
    ("feature_desc", "특징", "text"),
    ("memo", "비고", "text"),
]
FIELD_KIND = {k: v for k, _, v in FIELDS}
FIELD_LABEL = {k: lb for k, lb, _ in FIELDS}

# 열 이름 사전 — 중개사 매물장에서 실제로 쓰는 말. 긴 말이 먼저 걸려야 한다
# ('전용면적' 이 '면적' 보다, '소유자연락처' 가 '연락처' 보다 먼저).
HEADER_SYN: dict[str, list[str]] = {
    "trade_type": ["거래유형", "거래구분", "거래종류", "거래형태", "매매전세", "거래", "구분",
                   "매물구분", "종류", "tradetype", "trade", "type of deal"],
    "type": ["물건종류", "물건유형", "매물종류", "매물유형", "종별", "용도", "물건종별", "유형"],
    "complex_name": ["단지명", "아파트명", "건물명", "물건명", "단지", "아파트", "건물", "명칭",
                     "상호", "빌딩명", "complexname", "complex", "building", "name"],
    "address": ["소재지주소", "물건지주소", "지번주소", "도로명주소", "소재지", "물건지", "주소",
                "지번", "번지", "위치", "address", "addr", "location"],
    "dong": ["동수", "동번호", "건물동", "dong", "동"],
    "ho": ["호수", "호실", "호번호", "ho", "unit", "호"],
    "price": ["매매가격", "매매가", "매도가", "매가", "분양가", "희망가", "희망금액", "호가",
              "거래가", "매물가", "가액", "가격", "금액", "매매", "price", "sale", "amount"],
    "deposit": ["보증금액", "보증금", "전세보증금", "전세가격", "전세가", "임대보증금", "보증",
                "전세", "deposit", "jeonse"],
    "rent_price": ["월임대료", "월세금액", "월차임", "월세", "차임", "임대료", "월임대", "월",
                   "rent", "monthly"],
    "premium": ["권리금액", "권리금", "권리", "premium"],
    "maintenance_fee": ["관리비용", "월관리비", "관리비", "maintenance"],
    "loan_amount": ["융자금액", "융자금", "융자", "대출금", "대출", "근저당", "loan"],
    "area2_m2": ["전용면적", "전용㎡", "전용평", "전용", "실면적", "면적", "평형", "평수",
                 "넓이", "크기", "exclusive", "area", "size"],
    "area1_m2": ["공급면적", "분양면적", "계약면적", "공급", "분양평", "supply"],
    "land_area_m2": ["대지면적", "토지면적", "대지", "부지면적", "land"],
    "total_area_m2": ["연면적", "건물면적", "총면적", "totalarea"],
    "floor_info": ["해당층", "층수", "층/총층", "층", "floor"],
    "total_floor": ["총층수", "전체층", "총층", "건물층수", "totalfloor"],
    "direction": ["방향", "향", "direction"],
    "room_cnt": ["방개수", "방수", "룸수", "침실", "방", "room", "rooms"],
    "bath_cnt": ["욕실수", "화장실수", "욕실", "화장실", "bath"],
    "parking": ["주차대수", "주차가능", "주차", "parking"],
    "move_in": ["입주가능일", "입주가능", "입주일", "입주", "movein"],
    "settle_ymd": ["잔금일자", "잔금일", "잔금", "balance"],
    "approve_ymd": ["사용승인일", "사용승인", "준공일", "준공", "연식", "사용검사"],
    "owner_name": ["소유자명", "임대인명", "매도인명", "소유자", "소유주", "임대인", "매도인",
                   "집주인", "의뢰인", "owner"],
    "owner_tel": ["소유자연락처", "임대인연락처", "매도인연락처", "소유자전화", "집주인연락처",
                  "owner tel", "owner phone"],
    "contact": ["연락처", "전화번호", "휴대폰", "핸드폰", "전화", "폰", "번호", "tel", "phone",
                "contact", "mobile"],
    "manager": ["담당자", "담당", "manager", "agent"],
    "current_biz": ["현업종", "업종", "영업종목", "현재업종"],
    "land_category": ["지목", "지목현황"],
    "land_use": ["용도지역", "용도지구", "지역지구"],
    "tenant_until": ["임차만기", "계약만기", "만기", "임대만기"],
    "feature_desc": ["매물특징", "특징", "상세설명", "광고문구", "설명", "내용", "description"],
    "memo": ["특이사항", "비고", "메모", "참고", "note", "remark", "memo"],
}
# 한 칸에 동·호가 같이 오는 열 — '101동 1502호', '101-1502'
DONGHO_SYN = ["동호수", "동호", "호수동", "동/호", "동·호"]
# 매물이 아닌 줄 — 소계·합계·머리 반복
# 한 시트에서 읽는 최대 줄. 넘으면 잘렸다고 화면에 말한다 — 조용히 자르면
# '다 가져왔다'로 읽힌다(실측: 3,000줄 파일에서 1건이 소리 없이 사라졌다).
_MAX_ROWS = 20000
JUNK_ROW = re.compile(r"^\s*(소계|합계|총계|계|비고|이상|끝|합|total)\s*$")

_SP = re.compile(r"[\s ]+")
_PARENS = re.compile(r"[（(\[［].*?[)）\]］]")


def norm_head(s) -> str:
    """열 이름 정규화 — 공백·괄호·기호를 털어 낸다. '전용 면적(㎡)' → '전용면적'.

    앞에 붙는 번호도 뗀다('① 단지명'·'1. 거래'·'※비고'). 다만 **구분 기호가 뒤따를 때만**
    뗀다 — 그냥 숫자를 떼면 값 '103동' 이 '동' 이 되어 데이터 줄이 머리글로 읽힌다(실측).
    """
    t = _PARENS.sub("", str(s or ""))
    t = re.sub(r"^\s*(?:[①-⑳❶-❿⓵-⓾※★☆◆◇▶▷]+\s*|\d+\s*[.)\]]\s*)", "", t)
    t = _SP.sub("", t)
    t = re.sub(r"[·ㆍ/\\|,.\-_~*#:;'\"]+", "", t)
    t = re.sub(r"[㎡m²평억만원won]+$", "", t, flags=re.I) if len(t) > 2 else t
    return t.strip().lower()


# ── 값 파서 ──────────────────────────────────────────────────────────────────
TRADE_WORDS = [
    (r"단기임대|단기", "월세"), (r"반전세", "월세"),
    (r"매매|매도|분양|매각|^매$|sale", "매매"),
    (r"전세|jeonse|^전$", "전세"),
    (r"월세|월임대|임대차|사글세|^월$|rent|lease", "월세"),
    (r"임대", "월세"),
]
# 물건 종류 낱말 — **여기가 유일한 원본**이다. local_api 의 문장 판별기도 이걸 쓴다.
# 두 벌로 두면 같은 말을 서로 다르게 읽는다(실측: '건물' 이 한쪽에서만 잡혔다).
# scope 가 쓰임을 가른다 — "cell" 은 '물건종류' 칸처럼 낱말만 든 칸에서만 본다.
# '건물' 은 문장에 흔해('그 건물 2층 상가') 자유 문장에서 보면 오독한다.
PTYPE_WORDS = [
    (r"지식산업센터|지산|아파트형\s*공장", "지식산업센터", "both"),
    (r"상가주택|통건물|통임대|꼬마빌딩|빌딩", "건물", "both"),
    (r"^건물$", "건물", "cell"),
    (r"분양권|입주권", "분양권", "both"),
    (r"재개발|재건축", "재개발", "both"),
    (r"아파트|아파|apt", "아파트", "both"),
    (r"오피스텔|오피(?!스)", "오피스텔", "both"),
    (r"빌라|다세대|연립|타운하우스", "빌라", "both"),
    (r"다가구|단독주택|전원주택|단독", "단독", "both"),
    (r"원룸|투룸|고시원", "원룸", "both"),
    (r"상가|점포|근린생활|근생|매장", "상가", "both"),
    (r"사무실|사무소|오피스|office", "사무실", "both"),
    (r"토지|나대지|농지|임야|전답|필지|대지|땅", "토지", "both"),
    (r"공장|창고", "공장", "both"),
]
PTYPES = [t for _, t, _ in PTYPE_WORDS]
_NUM = re.compile(r"-?\d+(?:,\d{3})*(?:\.\d+)?")
# 사람이 '없음' 대신 찍어 두는 기호들 — 값으로 담으면 비고가 '-' 로 채워진다
_BLANKISH = {"-", "–", "—", "ㆍ", ".", "..", "…", "x", "X", "n/a", "na", "없음", "무"}
_TEL = re.compile(r"0\d{1,2}[-.\s]?\d{3,4}[-.\s]?\d{4}")


def cell_text(v) -> str:
    """엑셀 칸 → 문자열. 날짜·시각 칸이 그대로 오는 경우가 흔하다."""
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, _dtime):
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def to_trade(v) -> str | None:
    t = _SP.sub("", cell_text(v))
    if not t:
        return None
    for pat, out in TRADE_WORDS:
        if re.search(pat, t, re.I):
            return out
    return None


def to_ptype(v, scope: str = "cell") -> str | None:
    """낱말 → 물건 종류. scope="text" 면 자유 문장에서 안전한 규칙만 쓴다."""
    t = _SP.sub("", cell_text(v))
    if not t:
        return None
    for pat, out, sc in PTYPE_WORDS:
        if sc != "both" and sc != scope:
            continue
        if re.search(pat, t, re.I):
            return out
    return None


def to_won(v, unit: str = "auto") -> int | None:
    """금액 → **원**. unit 은 열 단위로 정해 넘긴다(억|만|원|auto).

    '12억3,000' 처럼 글자가 붙어 있으면 글자가 이긴다 — 열 단위보다 그 칸이 정확하다.
    """
    t = _SP.sub("", cell_text(v)).replace(",", "")
    if not t or t in ("-", "0"):
        return None
    m = re.search(r"(\d+(?:\.\d+)?)\s*억(?:\s*(\d+(?:\.\d+)?)\s*(?:천만|천|만)?)?", t)
    if m:
        won = float(m.group(1)) * 1e8
        if m.group(2):
            tail = float(m.group(2))
            # '12억 3천' 은 3,000만이고 '12억 3,000' 도 3,000만이다
            won += tail * 1e7 if ("천" in t and tail < 10) else tail * 1e4
        return int(round(won))
    m = re.search(r"(\d+(?:\.\d+)?)\s*천만", t)
    if m:
        return int(round(float(m.group(1)) * 1e7))
    m = re.search(r"(\d+(?:\.\d+)?)\s*만", t)
    if m:
        return int(round(float(m.group(1)) * 1e4))
    m = _NUM.search(t)
    if not m:
        return None
    n = float(m.group(0).replace(",", ""))
    if not n:
        return None
    if unit == "억":
        return int(round(n * 1e8))
    if unit == "만":
        return int(round(n * 1e4))
    if unit == "원":
        return int(round(n))
    return int(round(n)) if n >= 1e8 else int(round(n * 1e4))


def to_area(v, unit: str = "㎡") -> float | None:
    """면적 → ㎡. '84.98/59.87' 처럼 공급/전용이 한 칸에 오면 **뒤엣것**(전용)을 쓴다."""
    t = _SP.sub("", cell_text(v))
    if not t:
        return None
    if "평" in t:
        m = _NUM.search(t)
        return round(float(m.group(0).replace(",", "")) * 3.3058, 2) if m else None
    nums = [float(x.replace(",", "")) for x in _NUM.findall(t)]
    if not nums:
        return None
    n = nums[-1] if len(nums) > 1 else nums[0]
    if not n:
        return None
    return round(n * 3.3058, 2) if unit == "평" else round(n, 2)


def to_floor(v) -> str | None:
    """층 → 매물장 표기('3/15'). 화면이 뒤에 '층'을 붙이므로 여기선 붙이지 않는다."""
    t = _SP.sub("", cell_text(v))
    if not t:
        return None
    t = t.replace("층", "")
    if re.fullmatch(r"[Bb지하]+\d+", t):
        return "B" + re.sub(r"\D", "", t)
    m = re.fullmatch(r"(-?\d+)\s*/\s*(\d+)", t)
    if m:
        return f"{m.group(1)}/{m.group(2)}"
    m = re.fullmatch(r"-?\d+", t)
    if m:
        return t
    return t[:12] or None


def to_int(v) -> int | None:
    m = _NUM.search(cell_text(v))
    if not m:
        return None
    try:
        return int(float(m.group(0).replace(",", "")))
    except ValueError:
        return None


def to_num(v) -> float | None:
    m = _NUM.search(cell_text(v))
    if not m:
        return None
    try:
        return float(m.group(0).replace(",", ""))
    except ValueError:
        return None


def _tel_digits(t: str) -> str:
    d = re.sub(r"\D", "", t)
    if len(d) == 10 and d.startswith("10"):
        d = "0" + d          # 엑셀이 숫자로 저장하며 앞의 0 을 지운 번호(1012345678)
    return d


def to_tel(v) -> str | None:
    t = cell_text(v)
    d = _tel_digits(t)
    if len(d) not in (9, 10, 11):
        m = _TEL.search(t)   # 글 속에 번호가 섞여 있을 때만 잘라 낸다
        d = _tel_digits(m.group(0)) if m else ""
    if len(d) < 9 or len(d) > 11:
        return None
    # 서울(02)만 지역번호가 두 자리고 나머지는 셋이다(031·051·064). 길이로만 나누면
    # '031-355-1234' 가 '03-1355-1234' 가 된다(실측).
    head = d[:2] if d.startswith("02") else d[:3]
    rest = d[len(head):]
    return f"{head}-{rest[:-4]}-{rest[-4:]}" if len(rest) > 4 else f"{head}-{rest}"


def to_ymd(v) -> str | None:
    """날짜 → YYYY-MM-DD. 못 읽으면 적힌 대로 둔다('즉시', '세입자 만기 후')."""
    if isinstance(v, (datetime, date)):
        return (v.date() if isinstance(v, datetime) else v).isoformat()
    t = cell_text(v)
    if not t:
        return None
    # 구분자가 있는 형태와 8자리를 따로 본다. 느슨하게 하나로 두면 '2026-11'(연-월)이
    # 2026-01-01 로 쪼개진다 — 정규식이 '11' 을 '1'+'1' 로 되짚기 때문이다(실측).
    m = (re.search(r"(\d{4})\s*[-./년]\s*(\d{1,2})\s*[-./월]\s*(\d{1,2})\s*일?", t)
         or re.search(r"(?<!\d)(\d{4})(\d{2})(\d{2})(?!\d)", t))
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 1900 <= y <= 2100 and 1 <= mo <= 12 and 1 <= d <= 31:
            return f"{y:04d}-{mo:02d}-{d:02d}"
    m = re.fullmatch(r"(\d{2})\D(\d{1,2})\D(\d{1,2})", t.strip())
    if m:
        return f"20{int(m.group(1)):02d}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    m = re.search(r"(\d{4})\D{0,3}(\d{1,2})\s*월?\s*$", t)   # '2026년 8월' 은 사이가 두 글자다
    if m and 1900 <= int(m.group(1)) <= 2100:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}"
    return t[:20]


# '101동 1502호' 는 동·호지만 '역삼동 823-21' 은 주소다. 둘을 가르는 것은 두 가지다 —
# 건물 동은 숫자거나 한 글자(가동·A동)이고, 뒤에 '호'가 붙는다. 법정동은 두 글자 이상이고
# 뒤엣것은 지번이라 '호'가 없다. 그래서 주소를 훑을 땐 '호'를 반드시 요구한다.
_DONGHO = re.compile(r"(?:^|\s)(\d{1,4}|[A-Za-z]|[가-힣])\s*동\s*(\d{1,5})\s*호")
# 동·호 전용 칸이라면 '호'가 없어도 동·호로 읽는다 — 그 칸에 주소가 올 리 없다
_DONGHO_LOOSE = re.compile(r"(?:^|\s)(\d{1,4}|[A-Za-z]|[가-힣]{1,2})\s*동\s*(\d{1,5})\s*호?")
_DONGHO2 = re.compile(r"^\s*(\d{1,4})\s*[-/]\s*(\d{2,5})\s*$")


def split_dongho(v) -> tuple[str | None, str | None]:
    """'101동 1502호' · '101-1502' → ('101동', '1502호'). 못 가르면 (원문, None)."""
    t = cell_text(v)
    if not t:
        return None, None
    m = _DONGHO_LOOSE.search(t)
    if m:
        return f"{m.group(1)}동", f"{m.group(2)}호"
    m = _DONGHO2.match(t)
    if m:
        return f"{m.group(1)}동", f"{m.group(2)}호"
    return t[:20] or None, None


def tidy_unit(v, kind: str) -> str | None:
    """동·호 한 칸 — '101' → '101동' / '1502' → '1502호'. 이미 붙어 있으면 그대로."""
    t = _SP.sub("", cell_text(v))
    if not t:
        return None
    if kind == "dong":
        return t if t.endswith("동") else (f"{t}동" if re.fullmatch(r"[0-9A-Za-z가-힣]{1,4}", t) else t[:20])
    return t if t.endswith("호") else (f"{t}호" if re.fullmatch(r"\d{1,5}", t) else t[:20])


# ── 파일 읽기 ────────────────────────────────────────────────────────────────
def read_sheets(data: bytes, filename: str = "") -> list[dict]:
    """엑셀·CSV → [{name, rows}]. rows 는 칸 값 그대로(문자열 변환 전)."""
    name = (filename or "").lower()
    # 확장자보다 내용이 먼저다. '.csv' 로 저장된 엑셀 파일이 실제로 온다 —
    # 그걸 글자로 읽으면 NUL 이 섞여 통째로 터진다(실측).
    if data[:4] == b"PK\x03\x04":
        return _read_xlsx(data)
    if data[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        return _read_xls(data)
    if name.endswith(".csv") or name.endswith(".txt"):
        return [{"name": "CSV", "rows": _read_csv(data)}]
    if name.endswith(".xls"):
        return _read_xls(data)
    try:
        return _read_xlsx(data)
    except Exception:
        try:
            return _read_xls(data)
        except Exception:
            return [{"name": "CSV", "rows": _read_csv(data)}]


def _read_csv(data: bytes) -> list[list]:
    for enc in ("utf-8-sig", "cp949", "euc-kr", "utf-16", "latin-1"):
        try:
            text = data.decode(enc)
            break
        except (UnicodeDecodeError, LookupError):
            continue
    else:
        text = data.decode("utf-8", "replace")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
    except csv.Error:
        dialect = csv.excel
    return [list(r) for r in csv.reader(io.StringIO(text), dialect)]


def _read_xlsx(data: bytes) -> list[dict]:
    import openpyxl
    # 병합 범위를 알아야 해서 read_only 를 쓰지 않는다(그 모드는 병합 정보를 안 준다).
    # 아주 큰 파일만 read_only 로 떨어뜨린다 — 그땐 병합 칸이 빈 칸으로 남는다.
    big = len(data) > 12 * 1024 * 1024
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=big)
    out = []
    for ws in wb.worksheets:
        if ws.sheet_state != "visible":
            continue
        rows = [list(r) for r in ws.iter_rows(values_only=True, max_row=_MAX_ROWS)]
        if not big:
            _unmerge(ws, rows)
        out.append({"name": ws.title, "rows": rows})
    wb.close()
    return out


def _unmerge(ws, rows: list) -> None:
    """병합된 칸의 값을 그 범위 전체에 펴 놓는다 — 사람 눈에 보이는 대로 만든다."""
    for rng in list(getattr(ws, "merged_cells", []).ranges if hasattr(ws, "merged_cells") else []):
        r0, c0, r1, c1 = rng.min_row, rng.min_col, rng.max_row, rng.max_col
        if r0 - 1 >= len(rows) or c0 - 1 >= len(rows[r0 - 1] or []):
            continue
        v = rows[r0 - 1][c0 - 1]
        if v is None:
            continue
        for i in range(r0 - 1, min(r1, len(rows))):
            row = rows[i]
            for j in range(c0 - 1, min(c1, len(row))):
                if row[j] is None:
                    row[j] = v


def _read_xls(data: bytes) -> list[dict]:
    import xlrd
    bk = xlrd.open_workbook(file_contents=data)
    out = []
    for sh in bk.sheets():
        rows = []
        for i in range(min(sh.nrows, _MAX_ROWS)):
            r = []
            for j in range(sh.ncols):
                c = sh.cell(i, j)
                if c.ctype == xlrd.XL_CELL_DATE:
                    try:
                        r.append(datetime(*xlrd.xldate_as_tuple(c.value, bk.datemode)))
                        continue
                    except Exception:
                        pass
                r.append(c.value)
            rows.append(r)
        out.append({"name": sh.name, "rows": rows})
    return out


# ── 머리글 찾기 ──────────────────────────────────────────────────────────────
def _head_score(row) -> tuple[int, dict]:
    """이 줄이 머리글이면 몇 칸이나 우리 칸에 걸리나."""
    hit, seen = 0, {}
    for j, c in enumerate(row):
        f = match_header(c)
        if f and f not in seen:
            seen[f] = j
            hit += 1
    return hit, seen


def match_header(cell) -> str | None:
    """열 이름 한 칸 → 우리 칸 키. 긴 말부터 본다."""
    h = norm_head(cell)
    if not h or len(h) > 24:
        return None
    for syn in DONGHO_SYN:
        if h == norm_head(syn):
            return "_dongho"
    best = None
    for field, words in HEADER_SYN.items():
        for w in words:
            wn = norm_head(w)
            if not wn:
                continue
            if h == wn:
                return field
            if len(wn) >= 2 and wn in h and (best is None or len(wn) > best[1]):
                best = (field, len(wn))
    for syn in DONGHO_SYN:
        wn = norm_head(syn)
        if len(wn) >= 2 and wn in h and (best is None or len(wn) > best[1]):
            best = ("_dongho", len(wn))
    return best[0] if best else None


def _looks_header(row) -> tuple[bool, dict]:
    """이 줄이 진짜 머리글인가.

    '목련아파트 | 매매 | 62,000' 같은 데이터 줄도 낱말 두 개가 열 이름과 겹친다 —
    그것만 보고 머리글로 판정하면 첫 줄부터 읽기를 멈춘다(실측). 머리글은
    ① 채워진 칸 대부분이 열 이름이고 ② 숫자만 든 칸이 없다.
    """
    filled = [c for c in row if cell_text(c)]
    if len(filled) < 2:
        return False, {}
    for c in filled:
        t = _SP.sub("", cell_text(c)).replace(",", "")
        if _NUM.fullmatch(t):
            return False, {}
    hit, seen = _head_score(row)
    return (hit >= 2 and hit >= len(filled) * 0.6), seen


def find_header(rows: list) -> tuple[int, dict]:
    """머리글 줄 번호와 {칸키: 열번호}. 못 찾으면 (-1, {}).

    제목 줄·빈 줄이 위에 붙은 매물장이 흔해 위에서부터 20줄을 훑는다.
    """
    best = (-1, {}, 0)
    for i, row in enumerate(rows[:20]):
        if not row:
            continue
        hit, seen = _head_score(row)
        if hit > best[2]:
            best = (i, seen, hit)
    if best[2] < 2:
        return -1, {}
    i = best[0]
    # 머리글이 두 줄인 매물장 — 위는 '면적'·'금액' 같은 묶음 이름이고 아래가 '공급·전용',
    # '매매·전세' 다. 위 줄은 병합돼 있어 같은 말이 겹쳐 보인다 — 그게 신호다.
    labels = [norm_head(c) for c in rows[i] if norm_head(c)]
    if i + 1 < len(rows) and len(labels) != len(set(labels)):
        h2, seen2 = _head_score(rows[i + 1])
        if h2 >= 2:
            merged = dict(seen2)
            for f, j in best[1].items():        # 아래 줄이 비운 칸만 위 줄에서 가져온다
                if f not in merged and j not in merged.values():
                    merged[f] = j
            return i + 1, merged
    return i, best[1]


# ── 값으로 열 맞추기 ─────────────────────────────────────────────────────────
def infer_by_value(col_vals: list) -> str | None:
    """이름으로 못 맞춘 열 — 값의 생김새로 무엇인지 본다."""
    vals = [cell_text(v) for v in col_vals if cell_text(v)]
    if len(vals) < 2:
        return None
    n = len(vals)
    if sum(1 for v in vals if _TEL.search(v) or re.fullmatch(r"01\d{8,9}", re.sub(r"\D", "", v))) >= n * 0.6:
        return "contact"
    if sum(1 for v in vals if to_trade(v)) >= n * 0.8:
        return "trade_type"
    if sum(1 for v in vals if to_ptype(v)) >= n * 0.8:
        return "type"
    if sum(1 for v in vals if re.fullmatch(r"\d{1,3}\s*/\s*\d{1,3}", _SP.sub("", v))) >= n * 0.6:
        return "floor_info"
    if sum(1 for v in vals if _DONGHO.search(v) or _DONGHO2.match(v)) >= n * 0.6:
        return "_dongho"
    if sum(1 for v in vals if re.search(r"(동|읍|면|리)\s*\d+(-\d+)?$|[시도]\s", v)) >= n * 0.6:
        return "address"
    if sum(1 for v in vals if re.fullmatch(r"(남향|북향|동향|서향|남동|남서|북동|북서)", _SP.sub("", v))) >= n * 0.6:
        return "direction"
    return None


# ── 열 단위 ──────────────────────────────────────────────────────────────────
def money_unit(header, vals: list, field: str = "") -> str:
    """이 금액 열을 억으로 읽을지 만원으로 읽을지 원으로 읽을지.

    같은 '85,000' 이 매매가 열에서는 8억 5천이고, '12' 는 12억이다. 열의 값 분포를
    보고 정하되 머리글에 단위가 적혀 있으면 그게 이긴다.
    """
    h = str(header or "")
    if re.search(r"억", h):
        return "억"
    if re.search(r"만\s*원|만원|\(만\)", h):
        return "만"
    if re.search(r"원(?!룸)", h) and not re.search(r"만", h):
        return "원"
    nums = []
    for v in vals:
        t = _SP.sub("", cell_text(v)).replace(",", "")
        if not t or re.search(r"억|만|천", t):
            continue
        m = _NUM.search(t)
        if m:
            try:
                f = float(m.group(0))
            except ValueError:
                continue
            if f > 0:
                nums.append(f)
    if not nums:
        return "auto"
    nums.sort()
    mid = nums[len(nums) // 2]
    if mid >= 1e8:
        return "원"
    # 월세·관리비를 억으로 읽으면 45만원이 45억이 된다(실측). 작은 수가 정상인 칸이다
    if mid < 100 and field not in ("rent_price", "maintenance_fee"):
        return "억"
    return "만"


def area_unit(header, vals: list) -> str:
    h = str(header or "")
    if re.search(r"평", h) and not re.search(r"평형", h):
        return "평"
    if re.search(r"㎡|m2|m²|제곱", h, re.I):
        return "㎡"
    nums = [to_num(v) for v in vals]
    nums = [x for x in nums if x and x > 0]
    if not nums:
        return "㎡"
    nums.sort()
    mid = nums[len(nums) // 2]
    # 평은 대개 10~90, ㎡ 는 30~300. 겹치는 구간이 있어 머리글이 없으면 ㎡ 로 둔다
    return "평" if mid < 15 else "㎡"


# ── 분석 ─────────────────────────────────────────────────────────────────────
def analyze(data: bytes, filename: str = "", max_rows: int = _MAX_ROWS,
            override: dict | None = None) -> dict:
    """엑셀 → {sheets:[{name, header_row, columns, rows, skipped}]}. 저장하지 않는다.

    override 는 사람이 화면에서 고친 열 매핑이다 — {시트명: {"fields": {열번호: 칸키},
    "units": {열번호: "억"}}}. 자동 인식이 틀렸을 때 그 자리에서 바꾸지 못하면
    파일 자체를 못 쓰게 된다. 고친 뒤에는 **다시 읽어야** 값도 따라 바뀐다.
    """
    sheets = read_sheets(data, filename)
    out = []
    for sh in sheets:
        rows = [r for r in sh["rows"] if r is not None]
        if not rows:
            continue
        res = _analyze_sheet(sh["name"], rows, max_rows, (override or {}).get(sh["name"]) or {})
        if res:
            out.append(res)
    if not out:
        raise ValueError("읽을 수 있는 표가 없어요. 첫 시트에 매물 목록이 있는지 확인해 주세요.")
    out.sort(key=lambda x: (x["header_row"] > 0, len(x["columns"]), len(x["rows"])), reverse=True)
    return {"sheets": out, "filename": filename}


def _analyze_sheet(name: str, rows: list, max_rows: int, ovr: dict | None = None) -> dict | None:
    hidx, byname = find_header(rows)
    header = rows[hidx] if hidx >= 0 else []
    body = rows[hidx + 1:] if hidx >= 0 else rows
    while body and not any(cell_text(c) for c in body[-1]):
        body.pop()                      # 엑셀이 습관적으로 남기는 꼬리 빈 줄은 세지 않는다
    cut = max(0, len(body) - max_rows)
    body = body[:max_rows]
    if not body:
        return None
    width = max([len(r) for r in body] + [len(header)])

    def col(j):
        return [r[j] if j < len(r) else None for r in body[:60]]

    # ① 이름으로 맞춘 열
    mapping: dict[int, str] = {j: f for f, j in byname.items()}
    # ② 남은 열은 값으로
    inferred: set = set()
    for j in range(width):
        if j in mapping:
            continue
        got = infer_by_value(col(j))
        if got and got not in mapping.values():
            mapping[j] = got
            inferred.add(j)
    _infer_leftovers(mapping, width, col, header)
    # 사람이 고친 것이 마지막이다 — 자동 인식보다 언제나 우선한다
    forced = {int(k) for k in (ovr or {}).get("fields", {})}
    for k, v in ((ovr or {}).get("fields") or {}).items():
        j = int(k)
        if v:
            mapping[j] = v
        else:
            mapping.pop(j, None)      # '안 가져옴'
    # 매물 표가 아닌 시트를 걸러 낸다 — 안내·양식 시트가 앞에 붙은 매물장이 흔하다.
    # 어디 물건인지(단지·주소·동호)와 매물다운 값(금액·면적·거래) 이 둘 다 있어야 표로 본다.
    got = set(mapping.values())
    where = got & {"complex_name", "address", "dong", "ho", "_dongho"}
    what = got & {"price", "deposit", "rent_price", "premium", "area2_m2", "area1_m2", "trade_type"}
    if not forced and (len(mapping) < 2 or not where or not what):
        return None

    columns = []
    units: dict[int, str] = {}
    # 못 알아본 열도 값이 있으면 함께 세운다. 목록에 없으면 사람이 고를 수가 없어
    # 그 열은 영영 못 가져온다 — 자동 인식의 빈틈을 사람이 메울 길을 남긴다.
    shown = sorted(set(mapping) | {j for j in range(width)
                                   if any(cell_text(v) for v in col(j))})
    for j in shown:
        f = mapping.get(j, "")
        kind = "unit" if f == "_dongho" else FIELD_KIND.get(f, "text" if f else "")
        u = ""
        forced_u = ((ovr or {}).get("units") or {}).get(str(j))
        if kind == "money":
            u = forced_u or money_unit(header[j] if j < len(header) else "", col(j), f)
            units[j] = u
        elif kind == "area":
            u = forced_u or area_unit(header[j] if j < len(header) else "", col(j))
            units[j] = u
        columns.append({
            "index": j,
            "header": cell_text(header[j]) if j < len(header) else "",
            "field": f,
            "label": "동·호" if f == "_dongho" else FIELD_LABEL.get(f, f),
            "unit": u,
            "by": ("고침" if j in forced else "" if not f
                   else "이름" if j in byname.values()
                   else "값" if j in inferred else "짐작"),
            "kind": kind,
            "sample": [cell_text(v) for v in col(j)[:3] if cell_text(v)][:3],
        })

    parsed, skipped, stopped = [], [], 0
    mine = set(mapping.values())
    for i, r in enumerate(body):
        # 한 시트에 표가 둘인 매물장이 있다('[매매 물건]' 아래에 '[전월세 물건]').
        # 그때 두 번째 표는 열 순서가 다른데 첫 표의 열로 읽으면 보증금이 매매가로,
        # 월세가 면적으로 들어간다(실측). 머리글이 다시 나오면서 **열 구성이 달라지면**
        # 거기서 멈추고 그 사실을 말한다 — 틀린 값을 담는 것보다 낫다.
        is_head, seen = _looks_header(r)
        if is_head and set(seen) != mine:
            stopped = hidx + 2 + i
            break
        rec = _parse_row(r, mapping, units)
        why = _why_skip(rec, r)
        if why:
            skipped.append({"row": hidx + 2 + i, "why": why,
                            "text": " | ".join(cell_text(c) for c in r if cell_text(c))[:80]})
            continue
        rec["_row"] = hidx + 2 + i
        parsed.append(rec)
    return {"name": name, "header_row": hidx + 1 if hidx >= 0 else 0,
            "columns": columns, "rows": parsed, "skipped": skipped[:40],
            "n_skipped": len(skipped), "cut": cut, "stopped_at": stopped}


def _infer_leftovers(mapping: dict, width: int, col, header) -> None:
    """이름으로도 생김새로도 못 맞춘 열의 마지막 처리.

    머리글이 아예 없는 매물장이 실제로 온다. 그때 남는 것은 '글자 열'과 '숫자 열'인데,
    글자 열의 첫 번째는 대개 물건 이름이고 숫자 열은 금액이다. 확실하진 않으니
    이렇게 맞춘 열은 화면에서 사람이 바꿀 수 있어야 한다(by='짐작' 으로 표시한다).
    """
    taken = set(mapping.values())
    for j in range(width):
        if j in mapping:
            continue
        vals = [cell_text(v) for v in col(j) if cell_text(v)]
        if not vals:
            continue
        digits = sum(1 for v in vals if _NUM.fullmatch(v.replace(",", "")))
        if digits >= len(vals) * 0.8:
            n = to_won(vals[0], "auto") or 0
            # 값이 한 줄뿐인 열은 '매매가만 채운 매매 물건 한 건' 같은 경우다.
            # 그때도 액수가 물건 값이면 금액으로 본다 — 다만 짐작이라고 표시한다.
            if n >= 10_000_000 and "price" not in taken:
                mapping[j] = "price"; taken.add("price")
            elif n >= 1_000_000 and "deposit" not in taken:
                mapping[j] = "deposit"; taken.add("deposit")
            continue
        letters = sum(1 for v in vals if re.search(r"[가-힣A-Za-z]", v))
        if len(vals) >= 2 and letters >= len(vals) * 0.8 and max(len(v) for v in vals) <= 40:
            if "complex_name" not in taken and "address" not in taken:
                mapping[j] = "complex_name"; taken.add("complex_name")


def _parse_row(r: list, mapping: dict, units: dict) -> dict:
    rec: dict = {}
    for j, f in mapping.items():
        v = r[j] if j < len(r) else None
        if v is None or cell_text(v) == "" or cell_text(v).lower() in _BLANKISH:
            continue
        if f == "_dongho":
            d, h = split_dongho(v)
            if d and not rec.get("dong"):
                rec["dong"] = d
            if h and not rec.get("ho"):
                rec["ho"] = h
            continue
        kind = FIELD_KIND.get(f, "text")
        if kind == "money":
            got = to_won(v, units.get(j, "auto"))
        elif kind == "area":
            got = to_area(v, units.get(j, "㎡"))
        elif kind == "trade":
            got = to_trade(v)
        elif kind == "ptype":
            got = to_ptype(v)
        elif kind == "floor":
            got = to_floor(v)
        elif kind == "int":
            got = to_int(v)
        elif kind == "num":
            got = to_num(v)
        elif kind == "tel":
            got = to_tel(v)
        elif kind == "ymd":
            got = to_ymd(v)
        elif kind == "unit":
            got = tidy_unit(v, f)
        else:
            got = cell_text(v)[:300] or None
        if got not in (None, ""):
            rec[f] = got
    _post(rec)
    return rec


def _post(rec: dict) -> None:
    """줄 하나를 다 읽고 나서 정리 — 다른 칸을 같이 봐야 알 수 있는 것들."""
    # 거래유형이 없으면 금액 생김새로 정한다. 월세가 있으면 월세, 보증금만 있으면 전세
    if not rec.get("trade_type"):
        if rec.get("rent_price"):
            rec["trade_type"] = "월세"
        elif rec.get("deposit") and not rec.get("price"):
            rec["trade_type"] = "전세"
        elif rec.get("price"):
            rec["trade_type"] = "매매"
    # '금액' 한 칸으로 다 적는 매물장 — 전세·월세면 그건 매매가가 아니라 보증금이다
    if rec.get("trade_type") in ("전세", "월세") and rec.get("price") and not rec.get("deposit"):
        rec["deposit"] = rec.pop("price")
    # 주소에 동·호가 섞여 있으면 떼어 낸다
    if rec.get("address") and not rec.get("ho"):
        m = _DONGHO.search(rec["address"])
        if m:
            rec.setdefault("dong", f"{m.group(1)}동")
            rec["ho"] = f"{m.group(2)}호"
            rec["address"] = _SP.sub(" ", rec["address"][:m.start()]).strip() or rec["address"]
    # 단지명 칸에 주소가 통째로 들어온 매물장이 있다 — 주소가 비었으면 그리로 옮긴다
    if rec.get("complex_name") and not rec.get("address"):
        if re.search(r"(동|읍|면|리)\s*\d+(-\d+)?\s*$", rec["complex_name"]):
            rec["address"] = rec.pop("complex_name")
    # 층 하나만 있고 총층이 따로 있으면 '3/15' 로 합친다
    if rec.get("floor_info") and rec.get("total_floor") and "/" not in str(rec["floor_info"]):
        rec["floor_info"] = f"{rec['floor_info']}/{int(rec['total_floor'])}"
    # 물건 종류 — 칸이 말해 주는 것이 이름보다 정확하다.
    # '역삼빌딩'은 상가의 건물 **이름**이지 종류가 아니다(실측: 상가 매물장이 통째로
    # '건물'로 잡혔다). 그래서 이름에서 얻은 '건물'은 버린다.
    if not rec.get("type"):
        if rec.get("premium") or rec.get("current_biz"):
            rec["type"] = "상가"
        elif rec.get("land_category") or rec.get("land_use"):
            rec["type"] = "토지"
        elif (rec.get("land_area_m2") and not rec.get("area2_m2")
              and not rec.get("total_area_m2")):
            rec["type"] = "토지"      # 연면적이 있으면 건물이 서 있다 — 땅만 파는 게 아니다
    if not rec.get("type"):
        for k in ("feature_desc", "memo", "address", "complex_name"):
            got = to_ptype(rec.get(k))
            if got and not (k == "complex_name" and got == "건물"):
                rec["type"] = got
                break


def _why_skip(rec: dict, raw: list) -> str:
    """이 줄을 왜 못 담았는지. 조용히 버리면 몇 건이 사라졌는지도 모른다."""
    texts = [cell_text(c) for c in raw if cell_text(c)]
    if not texts:
        return "빈 줄"
    first = next((cell_text(c) for c in raw if cell_text(c)), "")
    if JUNK_ROW.match(first):
        return "매물 줄이 아님(소계·합계)"
    # 칸 하나에 글자만 있는 줄은 소제목이다('[전월세 물건]'). 매물이라면 금액이든
    # 동·호든 하나는 더 있다.
    if len(texts) == 1 and not any(rec.get(k) for k in ("price", "deposit", "rent_price",
                                                        "premium", "dong", "ho")):
        return "매물 줄이 아님(소제목)"
    if _looks_header(raw)[0] and len(texts) >= 3:
        return "머리글이 다시 나온 줄"
    has_where = any(rec.get(k) for k in ("complex_name", "address", "dong", "ho"))
    has_money = any(rec.get(k) for k in ("price", "deposit", "rent_price", "premium"))
    if not has_where and not has_money:
        return "매물로 볼 값이 없음"
    if not has_where:
        return "어디 물건인지 없음(단지·주소·동호가 다 빔)"
    return ""


def dedup_key(rec: dict) -> tuple:
    """같은 물건인지 보는 열쇠 — 한 파일 안의 중복과 이미 있는 매물을 함께 거른다.

    금액은 price 와 deposit 중 있는 것을 쓴다. 매물장은 전세·월세의 보증금을 price
    자리에 담기 때문이다 — 파서가 낸 deposit 만 보면 이미 담아 둔 전세를 못 알아본다.
    """
    where = _SP.sub("", str(rec.get("complex_name") or rec.get("address") or ""))
    money = int(rec.get("price") or rec.get("deposit") or 0)
    return (where, _SP.sub("", str(rec.get("dong") or "")), _SP.sub("", str(rec.get("ho") or "")),
            str(rec.get("trade_type") or ""), money, int(rec.get("rent_price") or 0))


if __name__ == "__main__":  # 손으로 확인할 때
    import sys
    with open(sys.argv[1], "rb") as fp:
        res = analyze(fp.read(), sys.argv[1])
    for sh in res["sheets"]:
        print(f"[{sh['name']}] 머리글 {sh['header_row']}행 · {len(sh['rows'])}건 "
              f"· 건너뜀 {sh['n_skipped']}")
        for c in sh["columns"]:
            print(f"   {c['index']:>2} {c['header'][:14]:<14} → {c['label']}"
                  f"{'(' + c['unit'] + ')' if c['unit'] else ''} [{c['by']}] {c['sample']}")
        print(json.dumps(sh["rows"][:3], ensure_ascii=False, indent=1))
