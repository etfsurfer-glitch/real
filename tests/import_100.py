#!/usr/bin/env python3
"""매물장 가져오기 — 100가지 서식 시험.

정답을 먼저 정하고(RECS), 그것을 100가지 서식으로 **뿌린 뒤 되읽어 대조**한다.
서식마다 기대값을 손으로 적으면 기대값 자체가 틀릴 수 있고, 무엇보다 100개를 손으로
못 적는다. 정답 → 서식 → 되읽기 로 닫으면 어떤 서식에서 무엇이 깨지는지가 바로 나온다.

    python3 tests/import_100.py                # 전부
    python3 tests/import_100.py 12 14          # 12·14번만(두 개씩 고치는 루틴용)
    python3 tests/import_100.py --from 40      # 40번부터
"""
from __future__ import annotations

import shutil
import sys
import tempfile
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "scripts"))
import listing_import as LI  # noqa: E402

OUT = Path(tempfile.mkdtemp(prefix="koczip-100-"))

# ── 정답 ────────────────────────────────────────────────────────────────────
# 금액은 원, 면적은 ㎡. 파서가 내놓아야 하는 값 그대로다.
A = {"complex_name": "래미안퍼스티지", "dong": "101동", "ho": "1502호", "trade_type": "매매",
     "price": 2850000000, "area2_m2": 84.97, "area1_m2": 114.9, "floor_info": "15/25",
     "contact": "010-1234-5678", "owner_name": "김소유", "type": "아파트",
     "settle_ymd": "2026-11-30", "memo": "올수리"}
B = {"complex_name": "아크로리버파크", "dong": "103동", "ho": "802호", "trade_type": "전세",
     "deposit": 1500000000, "area2_m2": 59.98, "area1_m2": 84.9, "floor_info": "8/35",
     "contact": "010-2222-3333", "owner_name": "박소유", "type": "아파트",
     "settle_ymd": "2027-01-15", "memo": "즉시입주"}
C = {"complex_name": "반포자이", "dong": "108동", "ho": "301호", "trade_type": "월세",
     "deposit": 100000000, "rent_price": 3500000, "area2_m2": 84.94, "area1_m2": 112.4,
     "floor_info": "3/32", "contact": "02-533-1234", "owner_name": "이소유", "type": "아파트",
     "settle_ymd": "2026-12-20", "memo": "반려동물 가능"}
RECS = [A, B, C]

# 비주거 정답 — 유형별 매물장에 쓴다
SG = {"address": "서울 강남구 역삼동 823-21", "complex_name": "역삼타워", "trade_type": "월세",
      "deposit": 50000000, "rent_price": 3500000, "premium": 80000000, "area2_m2": 49.5,
      "floor_info": "1/8", "current_biz": "카페", "type": "상가", "contact": "010-9999-8888"}
OF = {"address": "서울 강남구 논현동 12-3", "complex_name": "논현빌딩", "trade_type": "월세",
      "deposit": 30000000, "rent_price": 2000000, "area2_m2": 82.6, "floor_info": "5/12",
      "type": "사무실", "contact": "010-7777-6666"}
LD = {"address": "경기 양평군 서종면 문호리 산 12-3", "trade_type": "매매", "price": 850000000,
      "land_area_m2": 3305.0, "land_category": "임야", "land_use": "계획관리", "type": "토지"}
VL = {"address": "서울 강서구 화곡동 123-45", "ho": "301호", "trade_type": "전세",
      "deposit": 320000000, "area2_m2": 59.8, "floor_info": "3/5", "type": "빌라",
      "contact": "010-5555-4444"}
HS = {"address": "서울 은평구 응암동 55-2", "trade_type": "매매", "price": 1450000000,
      "land_area_m2": 198.3, "total_area_m2": 396.6, "type": "단독", "contact": "010-4444-3333"}
FT = {"address": "경기 화성시 향남읍 발안리 100-1", "trade_type": "매매", "price": 3200000000,
      "land_area_m2": 1652.0, "total_area_m2": 990.0, "type": "공장", "contact": "031-355-1234"}
KN = {"address": "경기 성남시 상대원동 5-1", "complex_name": "성남지산센터", "trade_type": "월세",
      "deposit": 50000000, "rent_price": 2500000, "area2_m2": 132.2, "floor_info": "7/15",
      "type": "지식산업센터", "contact": "010-3333-2222"}
OR = {"complex_name": "행복하우스", "ho": "201호", "trade_type": "월세", "deposit": 5000000,
      "rent_price": 450000, "area2_m2": 23.1, "floor_info": "2/5", "type": "원룸",
      "contact": "010-1212-3434"}
BD = {"address": "부산 해운대구 우동 1408", "complex_name": "해운대프라자", "trade_type": "매매",
      "price": 8500000000, "land_area_m2": 660.0, "total_area_m2": 2310.0, "type": "건물",
      "contact": "051-747-1234"}
BY = {"complex_name": "올림픽파크포레온", "dong": "302동", "ho": "1204호", "trade_type": "매매",
      "price": 1980000000, "area2_m2": 84.9, "type": "분양권", "contact": "010-8888-1111"}
RD = {"address": "서울 동작구 흑석동 90-1", "trade_type": "매매", "price": 1250000000,
      "land_area_m2": 99.2, "type": "재개발", "contact": "010-6666-5555"}

# ── 값 찍는 법 ───────────────────────────────────────────────────────────────
man = lambda v: f"{int(v) // 10000:,}" if v else ""                       # 85,000
eok = lambda v: (f"{v / 1e8:g}" if v else "")                              # 8.5
won = lambda v: f"{int(v):,}" if v else ""                                 # 850,000,000
eokman = lambda v: ("" if not v else                                       # 8억 5,000
                    (f"{int(v // 1e8)}억" + (f" {int(v % 1e8) // 10000:,}" if v % 1e8 else "")))
py = lambda v: (f"{round(v / 3.3058, 1)}" if v else "")                    # 25.7
m2 = lambda v: (f"{v}" if v else "")


def dh(r):                       # '101동 1502호'
    return " ".join(x for x in [r.get("dong"), r.get("ho")] if x)


def dh_dash(r):                  # '101-1502'
    d = (r.get("dong") or "").rstrip("동"); h = (r.get("ho") or "").rstrip("호")
    return f"{d}-{h}" if d and h else (h or d or "")


# 열 이름 → (기여하는 정답 칸, 값 찍는 함수)
COLS = {
    "단지명": ("complex_name", lambda r: r.get("complex_name", "")),
    "건물명": ("complex_name", lambda r: r.get("complex_name", "")),
    "아파트명": ("complex_name", lambda r: r.get("complex_name", "")),
    "물건명": ("complex_name", lambda r: r.get("complex_name", "")),
    "Complex": ("complex_name", lambda r: r.get("complex_name", "")),
    "소재지": ("address", lambda r: r.get("address", "")),
    "물건지": ("address", lambda r: r.get("address", "")),
    "주소": ("address", lambda r: r.get("address", "")),
    "Address": ("address", lambda r: r.get("address", "")),
    "동": ("dong", lambda r: (r.get("dong") or "").rstrip("동")),
    "호": ("ho", lambda r: (r.get("ho") or "").rstrip("호")),
    "동호수": (("dong", "ho"), dh),
    "동/호": (("dong", "ho"), dh_dash),
    "거래": ("trade_type", lambda r: r.get("trade_type", "")),
    "거래구분": ("trade_type", lambda r: r.get("trade_type", "")),
    "Trade": ("trade_type", lambda r: {"매매": "sale", "전세": "jeonse", "월세": "rent"}
              .get(r.get("trade_type"), "")),
    "물건종류": ("type", lambda r: r.get("type", "")),
    "매매가": ("price", lambda r: man(r.get("price"))),
    "매매가(억)": ("price", lambda r: eok(r.get("price"))),
    "매매가(원)": ("price", lambda r: won(r.get("price"))),
    "매매금액": ("price", lambda r: eokman(r.get("price"))),
    "보증금": ("deposit", lambda r: man(r.get("deposit"))),
    "전세가": ("deposit", lambda r: man(r.get("deposit"))),
    "월세": ("rent_price", lambda r: man(r.get("rent_price"))),
    "권리금": ("premium", lambda r: man(r.get("premium"))),
    "전용면적": ("area2_m2", lambda r: m2(r.get("area2_m2"))),
    "전용(평)": ("area2_m2", lambda r: py(r.get("area2_m2"))),
    "공급면적": ("area1_m2", lambda r: m2(r.get("area1_m2"))),
    "대지면적": ("land_area_m2", lambda r: m2(r.get("land_area_m2"))),
    "연면적": ("total_area_m2", lambda r: m2(r.get("total_area_m2"))),
    "층": ("floor_info", lambda r: r.get("floor_info", "")),
    "연락처": ("contact", lambda r: r.get("contact", "")),
    "소유자": ("owner_name", lambda r: r.get("owner_name", "")),
    "잔금일": ("settle_ymd", lambda r: r.get("settle_ymd", "")),
    "비고": ("memo", lambda r: r.get("memo", "")),
    "지목": ("land_category", lambda r: r.get("land_category", "")),
    "용도지역": ("land_use", lambda r: r.get("land_use", "")),
    "현업종": ("current_biz", lambda r: r.get("current_biz", "")),
}

CASES: list[dict] = []


def add(name, headers, recs=None, *, fmt="xlsx", pre=None, post=None, merge=None,
        sheets=None, tweak=None, skip_fields=(), note=""):
    """한 서식. headers 는 COLS 의 키 목록, pre/post 는 표 위·아래에 붙일 줄."""
    CASES.append({"name": name, "headers": list(headers), "recs": recs or RECS, "fmt": fmt,
                  "pre": pre or [], "post": post or [], "merge": merge or [],
                  "sheets": sheets, "tweak": tweak, "skip": set(skip_fields), "note": note})


def rows_of(headers, recs, tweak=None):
    out = [list(headers)]
    for r in recs:
        row = [COLS[h][1](r) for h in headers]
        out.append(tweak(row, r) if tweak else row)
    return out


def fields_of(headers, skip):
    got = set()
    for h in headers:
        k = COLS[h][0]
        got |= set(k) if isinstance(k, tuple) else {k}
    return got - set(skip)


# ═══════════════════════════════════════════════════════════════════════════
# 1~13 · 유형별 매물장 — 유형마다 열 구성이 다르다
add("01 아파트 표준", ["단지명", "동", "호", "거래", "매매가", "보증금", "월세", "전용면적", "층", "연락처"])
add("02 아파트 동호한칸", ["단지명", "동호수", "거래", "매매가", "보증금", "월세", "전용면적", "층"])
add("03 아파트 공급전용", ["단지명", "동", "호", "거래", "매매가", "보증금", "월세", "공급면적", "전용면적"])
add("04 상가", ["소재지", "건물명", "층", "전용면적", "보증금", "월세", "권리금", "현업종", "연락처"], [SG])
add("05 사무실", ["소재지", "건물명", "층", "전용면적", "보증금", "월세", "연락처"], [OF])
add("06 토지", ["소재지", "지목", "용도지역", "대지면적", "매매가", "물건종류"], [LD])
add("07 빌라", ["주소", "호", "거래", "보증금", "전용면적", "층", "연락처", "물건종류"], [VL])
add("08 단독", ["소재지", "거래", "매매가", "대지면적", "연면적", "연락처", "물건종류"], [HS])
add("09 공장", ["소재지", "거래", "매매가", "대지면적", "연면적", "연락처", "물건종류"], [FT])
add("10 지식산업센터", ["소재지", "건물명", "거래", "보증금", "월세", "전용면적", "층", "물건종류"], [KN])
add("11 원룸", ["건물명", "호", "거래", "보증금", "월세", "전용면적", "층", "연락처", "물건종류"], [OR])
add("12 건물", ["소재지", "건물명", "거래", "매매가", "대지면적", "연면적", "연락처", "물건종류"], [BD])
add("13 분양권·재개발", ["단지명", "동", "호", "거래", "매매가", "전용면적", "물건종류"], [BY])

# 14~25 · 머리글 위치와 구조
add("14 제목 1줄 위", ["단지명", "동", "호", "거래", "매매가", "전용면적"],
    pre=[["○○공인중개사사무소 매물장"]])
add("15 제목+빈줄", ["단지명", "동", "호", "거래", "매매가", "전용면적"],
    pre=[["매물장"], [], ["2026년 8월"]])
add("16 제목 5줄", ["단지명", "동", "호", "거래", "매매가", "전용면적"],
    pre=[["매물장"], ["작성자 김중개"], ["연락처 010-0000-0000"], [], ["※ 금액 단위 만원"]])
add("17 머리글 없음", ["단지명", "거래", "매매가", "연락처"],
    tweak=lambda row, r: row, skip_fields=("complex_name",),
    note="머리글 줄을 지운다")
CASES[-1]["drop_header"] = True
add("18 아래에 각주", ["단지명", "동", "호", "거래", "매매가", "전용면적"],
    post=[[], ["※ 금액은 만원 단위입니다"], ["작성 2026-08-08"]])
add("19 소계·합계", ["단지명", "동", "호", "거래", "매매가", "전용면적"],
    post=[["소계"], ["합계"]])
add("20 머리글 반복", ["단지명", "동", "호", "거래", "매매가", "전용면적"],
    post=[["단지명", "동", "호", "거래", "매매가", "전용면적"]])
add("21 중간 빈 줄", ["단지명", "동", "호", "거래", "매매가", "전용면적"],
    post=[[], [], []])
add("22 시트 여럿(안내 앞)", ["단지명", "동", "호", "거래", "매매가", "전용면적"],
    sheets=[("안내", [["작성요령"], ["금액은 만원 단위"]])])
add("23 시트 여럿(빈 시트 앞)", ["단지명", "동", "호", "거래", "매매가", "전용면적"],
    sheets=[("Sheet1", [[]])])
add("24 시트명이 숫자", ["단지명", "동", "호", "거래", "매매가", "전용면적"])
CASES[-1]["sheet_name"] = "2026"
add("25 한 줄짜리", ["단지명", "동", "호", "거래", "매매가", "전용면적"], [A])

# 26~40 · 단위
add("26 만원", ["단지명", "거래", "매매가", "보증금", "월세", "전용면적"])
add("27 억", ["단지명", "거래", "매매가(억)", "전용면적"], [A])
add("28 원", ["단지명", "거래", "매매가(원)", "전용면적"], [A])
add("29 억+만 섞임", ["단지명", "거래", "매매금액", "전용면적"], [A])
add("30 평 면적", ["단지명", "거래", "매매가", "전용(평)"], [A], skip_fields=("area2_m2",),
    note="평→㎡ 환산은 반올림 오차가 있어 값 대조에서 뺀다")
add("31 억(전세)", ["단지명", "거래", "전세가", "전용면적"], [B])
CASES[-1]["headers"] = ["단지명", "거래", "전세가", "전용면적"]
add("32 월세 작은 수", ["건물명", "호", "거래", "보증금", "월세", "전용면적"], [OR])
add("33 권리금 만원", ["소재지", "거래", "보증금", "월세", "권리금", "전용면적"], [SG])
add("34 금액에 '원' 붙음", ["단지명", "거래", "매매가", "전용면적"], [A],
    tweak=lambda row, r: [row[0], row[1], row[2] + "원", row[3]])
add("35 금액에 공백", ["단지명", "거래", "매매가", "전용면적"], [A],
    tweak=lambda row, r: [row[0], row[1], f" {row[2]} ", row[3]])
add("36 면적에 ㎡", ["단지명", "거래", "매매가", "전용면적"], [A],
    tweak=lambda row, r: [row[0], row[1], row[2], row[3] + "㎡"])
add("37 매매가 0/빈칸 섞임", ["단지명", "거래", "매매가", "보증금", "월세", "전용면적"])
add("38 '-' 로 빈칸 표시", ["단지명", "거래", "매매가", "보증금", "월세", "전용면적", "비고"],
    tweak=lambda row, r: [c if c else "-" for c in row])
add("39 대지·연면적 ㎡", ["소재지", "거래", "매매가", "대지면적", "연면적"], [HS])
add("40 전용/공급 한 칸", ["단지명", "거래", "매매가", "전용면적"], [A],
    tweak=lambda row, r: [row[0], row[1], row[2], f"{r['area1_m2']}/{r['area2_m2']}"])

# 41~52 · 동·호 표기
add("41 동/호 분리", ["단지명", "동", "호", "거래", "매매가"])
add("42 동호 한 칸(공백)", ["단지명", "동호수", "거래", "매매가"])
add("43 동호 한 칸(하이픈)", ["단지명", "동/호", "거래", "매매가"])
add("44 동에 '동' 붙음", ["단지명", "동", "호", "거래", "매매가"],
    tweak=lambda row, r: [row[0], r["dong"], r["ho"], row[3], row[4]])
add("45 호만 있음", ["건물명", "호", "거래", "보증금", "월세"], [OR])
add("46 주소에 동호 포함", ["소재지", "거래", "매매가"], [dict(A, address="서울 강남구 대치동 316 101동 1502호")],
    skip_fields=("address",))
add("47 동이 문자(가동)", ["단지명", "동", "호", "거래", "매매가"],
    [dict(A, dong="가동")], skip_fields=())
add("48 동이 A동", ["단지명", "동", "호", "거래", "매매가"], [dict(A, dong="A동")])
add("49 호가 4자리", ["단지명", "동", "호", "거래", "매매가"], [dict(A, ho="1502호")])
add("50 지하층", ["소재지", "거래", "보증금", "월세", "층"], [dict(SG, floor_info="B1")])
add("51 층 단독(총층 없음)", ["단지명", "거래", "매매가", "층"], [dict(A, floor_info="15")])
add("52 층에 '층' 붙음", ["단지명", "거래", "매매가", "층"], [dict(A, floor_info="15/25")],
    tweak=lambda row, r: [row[0], row[1], row[2], row[3] + "층"])

# 53~64 · 파일 형식·인코딩
add("53 CSV utf-8", ["단지명", "동", "호", "거래", "매매가", "전용면적"], fmt="csv-utf8")
add("54 CSV cp949", ["단지명", "동", "호", "거래", "매매가", "전용면적"], fmt="csv-cp949")
add("55 CSV euc-kr", ["단지명", "동", "호", "거래", "매매가", "전용면적"], fmt="csv-euckr")
add("56 CSV BOM", ["단지명", "동", "호", "거래", "매매가", "전용면적"], fmt="csv-bom")
add("57 TSV", ["단지명", "동", "호", "거래", "매매가", "전용면적"], fmt="tsv")
add("58 세미콜론", ["단지명", "동", "호", "거래", "매매가", "전용면적"], fmt="semi")
add("59 구형 xls", ["단지명", "동", "호", "거래", "매매가", "전용면적"], fmt="xls")
add("60 구형 xls 상가", ["소재지", "층", "보증금", "월세", "권리금", "현업종"], [SG], fmt="xls")
add("61 xlsx 확장자 없음", ["단지명", "동", "호", "거래", "매매가"], fmt="noext")
add("62 CSV 확장자인데 xlsx", ["단지명", "동", "호", "거래", "매매가"], fmt="xlsx-as-csv")
add("63 큰 파일 1000줄", ["단지명", "동", "호", "거래", "매매가", "전용면적"], fmt="xlsx")
CASES[-1]["repeat"] = 334
add("64 아주 큰 파일 5000줄", ["단지명", "동", "호", "거래", "매매가", "전용면적"], fmt="xlsx")
CASES[-1]["repeat"] = 1667

# 65~76 · 열 이름 변형
add("65 영문 머리글", ["Complex", "동", "호", "Trade", "매매가", "전용면적"])
add("66 영문 주소", ["Address", "Trade", "매매가", "전용면적"], [SG], skip_fields=("trade_type",))
add("67 머리글에 공백", ["단지명", "동", "호", "거래", "매매가", "전용면적"],
    tweak=None)
CASES[-1]["head_tweak"] = lambda h: f" {h} "
add("68 머리글에 줄바꿈", ["단지명", "동", "호", "거래", "매매가", "전용면적"])
CASES[-1]["head_tweak"] = lambda h: h.replace("면적", "\n면적")
add("69 머리글에 번호", ["단지명", "동", "호", "거래", "매매가", "전용면적"])
CASES[-1]["head_tweak"] = lambda h: f"① {h}"
add("70 머리글 괄호단위", ["단지명", "거래", "매매가", "전용면적"], [A])
CASES[-1]["head_tweak"] = lambda h: {"매매가": "매매가(만원)", "전용면적": "전용면적(㎡)"}.get(h, h)
add("71 머리글 전각", ["단지명", "동", "호", "거래", "매매가", "전용면적"])
CASES[-1]["head_tweak"] = lambda h: h.replace("(", "（").replace(")", "）")
add("72 낯선 머리글 섞임", ["단지명", "동", "호", "거래", "매매가", "전용면적"],
    note="정체불명 열이 끼어도 나머지는 읽는다")
CASES[-1]["extra_cols"] = [("코드", lambda r: "A-1024"), ("등급", lambda r: "A")]
add("73 우리가 안 쓰는 열 다수", ["단지명", "동", "호", "거래", "매매가", "전용면적"])
CASES[-1]["extra_cols"] = [(f"기타{i}", lambda r, i=i: f"v{i}") for i in range(1, 16)]
add("74 열 순서 뒤섞임", ["전용면적", "매매가", "거래", "호", "동", "단지명"])
add("75 빈 열이 사이사이", ["단지명", "동", "호", "거래", "매매가", "전용면적"])
CASES[-1]["blank_cols"] = True
add("76 머리글만 있고 값 없는 열", ["단지명", "동", "호", "거래", "매매가", "전용면적", "월세"])

# 77~88 · 엑셀 기능
# 병합하면 아래 칸의 값은 사라진다 — 같은 단지의 여러 호를 묶는 것이 병합의 실제 쓰임이다
add("77 병합 셀(단지명)", ["단지명", "동", "호", "거래", "매매가", "전용면적"],
    [A, dict(A, ho="1503호"), dict(A, ho="1504호")], merge=[("A2:A4", 0)])
add("78 두 줄 머리글", ["단지명", "공급면적", "전용면적", "매매가", "전용면적"], [A])
CASES[-1]["two_row_head"] = (["단지", "면적", "면적", "금액"], ["", "공급", "전용", "매매"])
CASES[-1]["headers"] = ["단지명", "공급면적", "전용면적", "매매가"]
add("79 수식 열", ["단지명", "동", "호", "거래", "매매가", "전용면적"])
CASES[-1]["extra_cols"] = [("평수", lambda r: "=1/1")]
add("80 통화 서식", ["단지명", "거래", "매매가(원)", "전용면적"], [A], fmt="currency")
add("81 숨긴 시트 앞", ["단지명", "동", "호", "거래", "매매가", "전용면적"],
    sheets=[("작업용", [["임시", "메모"], ["쓰지않음", "삭제"]])])
CASES[-1]["hide_first"] = True
add("82 틀 고정·필터", ["단지명", "동", "호", "거래", "매매가", "전용면적"], fmt="freeze")
add("83 숫자로 저장된 전화", ["단지명", "거래", "매매가", "연락처"], [A], fmt="telnum")
add("84 날짜 칸", ["단지명", "거래", "매매가", "잔금일"], [A], fmt="datecell")
add("85 날짜 문자열", ["단지명", "거래", "매매가", "잔금일"], [A])
add("86 줄바꿈 든 비고", ["단지명", "거래", "매매가", "비고"], [dict(A, memo="올수리\n남향\n즉시입주")])
add("87 앞뒤 공백 값", ["단지명", "거래", "매매가", "전용면적"], [A],
    tweak=lambda row, r: [f"  {c}  " for c in row])
add("88 아주 긴 비고", ["단지명", "거래", "매매가", "비고"], [dict(A, memo="가" * 400)],
    skip_fields=("memo",))

# 89~100 · 실무 잡음
add("89 중복 줄", ["단지명", "동", "호", "거래", "매매가"], [A, A, B],
    note="같은 물건이 두 번 — 읽기는 하고 중복 표시는 저장 때 한다")
add("90 어디 물건인지 없는 줄", ["단지명", "동", "호", "거래", "매매가"])
CASES[-1]["post"] = [["", "", "", "매매", "68,000"]]
add("91 금액 없는 줄", ["단지명", "동", "호", "거래", "매매가"])
CASES[-1]["post"] = [["미정단지", "", "", "", ""]]
add("92 전각 숫자", ["단지명", "거래", "매매가", "전용면적"], [A],
    tweak=lambda row, r: [row[0], row[1], row[2].translate(str.maketrans("0123456789", "０１２３４５６７８９")), row[3]])
add("93 거래유형 약어", ["단지명", "동", "호", "거래", "매매가", "보증금", "월세"],
    tweak=lambda row, r: [row[0], row[1], row[2], {"매매": "매", "전세": "전", "월세": "월"}[r["trade_type"]],
                          row[4], row[5], row[6]])
add("94 거래유형 없음", ["단지명", "동", "호", "매매가", "보증금", "월세"],
    note="금액 생김새로 거래유형을 정한다")
add("95 유형 낱말이 단지명에", ["단지명", "거래", "보증금", "전용면적"],
    [dict(VL, complex_name="화곡동 신축빌라")], skip_fields=("address",))
add("96 소유자·연락처 분리", ["단지명", "동", "호", "거래", "매매가", "소유자", "연락처"])
add("97 잔금·비고까지", ["단지명", "동", "호", "거래", "매매가", "보증금", "월세", "잔금일", "비고"])
add("98 상가+아파트 섞인 표", ["소재지", "단지명", "거래", "매매가", "보증금", "월세", "권리금", "전용면적"],
    [A, SG], skip_fields=("type",))
add("99 유형 열 있음", ["물건종류", "단지명", "소재지", "거래", "매매가", "보증금", "월세", "전용면적"],
    [A, SG, LD, VL], skip_fields=("area2_m2", "land_area_m2"))
add("100 전 유형 한 표", ["물건종류", "소재지", "단지명", "동", "호", "거래", "매매가", "보증금",
                     "월세", "권리금", "전용면적", "대지면적", "연면적", "층", "연락처"],
    [A, B, C, SG, OF, LD, VL, HS, FT, KN, OR, BD, BY, RD])


# ── 파일 굽기 ────────────────────────────────────────────────────────────────
def build(case, idx) -> Path:
    heads = list(case["headers"])
    ht = case.get("head_tweak")
    recs = case["recs"] * (case.get("repeat") or 1)
    body = rows_of(heads, recs, case.get("tweak"))
    header, data = body[0], body[1:]
    if ht:
        header = [ht(h) for h in header]
    for name, fn in (case.get("extra_cols") or []):
        header.append(name)
        for i, r in enumerate(recs):
            data[i].append(fn(r))
    if case.get("blank_cols"):
        header = sum([[h, ""] for h in header], [])
        data = [sum([[c, ""] for c in row], []) for row in data]
    rows = list(case["pre"])
    if case.get("two_row_head"):
        rows += [list(case["two_row_head"][0]), list(case["two_row_head"][1])]
    elif not case.get("drop_header"):
        rows.append(header)
    rows += data + list(case["post"])

    fmt = case["fmt"]
    ext = {"csv-utf8": ".csv", "csv-cp949": ".csv", "csv-euckr": ".csv", "csv-bom": ".csv",
           "tsv": ".csv", "semi": ".csv", "xls": ".xls", "noext": "",
           "xlsx-as-csv": ".csv"}.get(fmt, ".xlsx")
    p = OUT / f"{idx:03d}{ext}"

    if fmt.startswith("csv") or fmt in ("tsv", "semi"):
        sep = "\t" if fmt == "tsv" else (";" if fmt == "semi" else ",")
        import csv as _csv
        import io as _io
        buf = _io.StringIO()
        w = _csv.writer(buf, delimiter=sep)
        for r in rows:
            w.writerow(["" if c is None else str(c) for c in r])
        text = buf.getvalue()
        enc = {"csv-cp949": "cp949", "csv-euckr": "euc-kr"}.get(fmt, "utf-8")
        p.write_bytes(("﻿" + text if fmt == "csv-bom" else text).encode(enc, "replace"))
        return p
    if fmt == "xls":
        import xlwt
        bk = xlwt.Workbook(encoding="utf-8")
        sh = bk.add_sheet(case.get("sheet_name") or "매물")
        for i, r in enumerate(rows):
            for j, c in enumerate(r):
                if c not in (None, ""):
                    sh.write(i, j, c)
        bk.save(str(p))
        return p

    import openpyxl
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for nm, rr in (case.get("sheets") or []):
        ws0 = wb.create_sheet(title=nm[:31])
        for r in rr:
            ws0.append(list(r))
        if case.get("hide_first"):
            ws0.sheet_state = "hidden"
    ws = wb.create_sheet(title=(case.get("sheet_name") or "매물")[:31])
    for r in rows:
        ws.append(list(r))
    off = len(case["pre"])
    for rng, _c in case["merge"]:
        ws.merge_cells(rng)
    if fmt == "currency":
        for row in ws.iter_rows(min_row=off + 2):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '₩#,##0'
    if fmt == "freeze":
        ws.freeze_panes = f"A{off + 2}"
        ws.auto_filter.ref = f"A{off + 1}:F{off + 1 + len(data)}"
    if fmt == "telnum":
        for row in ws.iter_rows(min_row=off + 2):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.replace("-", "").isdigit() \
                        and cell.value.startswith("010"):
                    cell.value = int(cell.value.replace("-", ""))
    if fmt == "datecell":
        for row in ws.iter_rows(min_row=off + 2):
            for cell in row:
                if isinstance(cell.value, str) and len(cell.value) == 10 and cell.value[4] == "-":
                    try:
                        cell.value = datetime.strptime(cell.value, "%Y-%m-%d")
                    except ValueError:
                        pass
    wb.save(p)
    return p


def near(a, b) -> bool:
    if isinstance(b, float) and isinstance(a, (int, float)):
        return abs(float(a) - b) <= max(0.06, abs(b) * 0.006)
    if isinstance(b, int) and isinstance(a, (int, float)):
        return abs(float(a) - b) <= max(1.0, abs(b) * 0.006)
    return str(a or "") == str(b or "")


def check(case, sh) -> list:
    errs = []
    recs = case["recs"] * (case.get("repeat") or 1)
    want_n = len(recs)
    if len(sh["rows"]) != want_n:
        errs.append(f"건수 {len(sh['rows'])} ≠ {want_n}"
                    f" · 건너뜀 {[s['why'] for s in sh['skipped']][:3]}"
                    + (f" · {sh.get('stopped_at')}행에서 멈춤" if sh.get("stopped_at") else ""))
    fields = fields_of(case["headers"], case["skip"])
    for i, want in enumerate(recs[:12]):
        if i >= len(sh["rows"]):
            break
        got = sh["rows"][i]
        for f in sorted(fields):
            if f not in want or want.get(f) in (None, ""):
                continue
            if not near(got.get(f), want[f]):
                errs.append(f"[{i}] {f}: {got.get(f)!r} ≠ {want[f]!r}")
    return errs[:6]


def main(only: set | None, start: int) -> int:
    ok, bad = 0, []
    import time
    for idx, case in enumerate(CASES, 1):
        if only and idx not in only:
            continue
        if idx < start:
            continue
        p = build(case, idx)
        t = time.time()
        try:
            res = LI.analyze(p.read_bytes(), p.name)
            sh = res["sheets"][0]
            errs = check(case, sh)
        except Exception as e:  # noqa: BLE001
            errs = [f"터짐: {type(e).__name__} {e}"]
            sh = {"rows": []}
        el = (time.time() - t) * 1000
        if errs:
            bad.append((case["name"], errs))
            print(f"  ✗ {case['name']:<26} {len(sh['rows']):>5}건 {el:>6.0f}ms")
            for e in errs:
                print(f"       {e}")
        else:
            ok += 1
            print(f"  ✓ {case['name']:<26} {len(sh['rows']):>5}건 {el:>6.0f}ms")
    n = ok + len(bad)
    print(f"\n{n}가지 중 통과 {ok} ({ok / max(1, n) * 100:.0f}%)")
    if bad:
        print("실패:", ", ".join(x[0].split()[0] for x in bad))
    return len(bad)


if __name__ == "__main__":
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    start = 1
    if "--from" in sys.argv:
        start = int(sys.argv[sys.argv.index("--from") + 1])
        args = [a for a in args if a != str(start)]
    only = {int(a) for a in args if a.isdigit()} or None
    rc = main(only, start)
    shutil.rmtree(OUT, ignore_errors=True)
    raise SystemExit(min(1, rc))
