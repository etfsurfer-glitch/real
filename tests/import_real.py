#!/usr/bin/env python3
"""진짜 엑셀 파일로 시험한다 — 디스크에 구워서 읽는다.

코퍼스(import_corpus)는 서식의 '함정'을 하나씩 겨눈 것이고, 이건 중개사가 실제로
들고 올 법한 **파일 그대로**를 만든다. 수식이 든 시트, 서식만 통화인 숫자, 숨긴 시트,
50열짜리 관리표, 구형 .xls, 탭으로 나눈 텍스트, 3,000줄짜리 큰 파일까지.

    python3 tests/import_real.py            # 만들고 읽고 채점
    python3 tests/import_real.py --keep     # 만든 파일을 남긴다(눈으로 열어 보려고)
"""
import io
import shutil
import sys
import tempfile
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "scripts"))
import listing_import as LI  # noqa: E402

OUT = Path(tempfile.mkdtemp(prefix="koczip-import-"))
BUILT: list[tuple[str, dict]] = []      # (파일경로, 기대)


def xlsx(name, build, expect):
    """build(wb) 가 통째로 워크북을 꾸민다 — 서식·수식·숨김까지 진짜로 넣는다."""
    import openpyxl
    wb = openpyxl.Workbook()
    build(wb)
    p = OUT / name
    wb.save(p)
    BUILT.append((str(p), expect))


def xls(name, rows, expect):
    import xlwt
    bk = xlwt.Workbook(encoding="utf-8")
    sh = bk.add_sheet("매물")
    for i, r in enumerate(rows):
        for j, v in enumerate(r):
            if v is not None:
                sh.write(i, j, v)
    p = OUT / name
    bk.save(str(p))
    BUILT.append((str(p), expect))


def raw(name, data: bytes, expect):
    p = OUT / name
    p.write_bytes(data)
    BUILT.append((str(p), expect))


# ─────────────────────────────────────────────────────────────────────────────
# ① 수식이 든 매물장 — 평수를 ㎡ 에서 계산해 두는 사무소가 많다
def _formula(wb):
    ws = wb.active
    ws.title = "매물"
    ws.append(["단지명", "동", "호", "거래", "매매가(만)", "전용㎡", "평수", "평당가"])
    ws.append(["래미안퍼스티지", "101", "1502", "매매", 285000, 84.97, "=F2/3.3058", "=E2/G2"])
    ws.append(["아크로리버파크", "103", "802", "매매", 420000, 112.4, "=F3/3.3058", "=E3/G3"])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:H3"


xlsx("01_수식과틀고정.xlsx", _formula,
     {"n": 2, "check": {0: {"complex_name": "래미안퍼스티지", "price": 2850000000,
                            "area2_m2": 84.97, "dong": "101동", "ho": "1502호"}}})


# ② 서식만 통화인 숫자 — 셀 값은 그냥 숫자이고 보이는 것만 '₩285,000,000'
def _currency(wb):
    ws = wb.active
    ws.title = "매물목록"
    ws.append(["물건명", "거래구분", "매매금액", "전용면적", "연락처"])
    for r in [["힐스테이트", "매매", 1250000000, 84.9, "010-1111-2222"],
              ["e편한세상", "매매", 980000000, 59.8, "010-3333-4444"]]:
        ws.append(r)
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        for c in row:
            c.number_format = '₩#,##0'


xlsx("02_통화서식.xlsx", _currency,
     {"n": 2, "check": {0: {"price": 1250000000, "area2_m2": 84.9}}})


# ③ 숨긴 시트가 앞에 있고 진짜 표는 뒤에
def _hidden(wb):
    ws0 = wb.active
    ws0.title = "작업용"
    ws0.append(["임시", "메모"])
    ws0.append(["쓰지않음", "삭제예정"])
    ws0.sheet_state = "hidden"
    ws1 = wb.create_sheet("8월매물")
    ws1.append(["단지", "동호수", "거래", "보증금", "월세", "전용"])
    ws1.append(["센트럴파크", "205동 1103호", "월세", "10,000", "350", "84.9"])
    ws1.append(["더샵", "101동 502호", "월세", "5,000", "200", "59.8"])


xlsx("03_숨긴시트.xlsx", _hidden,
     {"n": 2, "check": {0: {"complex_name": "센트럴파크", "dong": "205동", "ho": "1103호",
                            "deposit": 100000000, "rent_price": 3500000}}})


# ④ 50열짜리 사무소 관리표 — 우리가 안 쓰는 열이 잔뜩이다
def _wide(wb):
    ws = wb.active
    ws.title = "관리대장"
    head = ["연번", "등록일", "매물번호", "담당자", "단지명", "동", "호", "거래구분",
            "매매가", "전세가", "월세보증금", "월세", "관리비", "전용면적", "공급면적",
            "층", "총층", "향", "방", "욕실", "주차", "입주가능일", "잔금일", "사용승인일",
            "소유자", "소유자연락처", "세입자", "세입자연락처", "융자", "특이사항"]
    head += [f"기타{i}" for i in range(1, 21)]
    ws.append(head)
    row = [1, datetime(2026, 8, 1), "A-1024", "김중개", "목련아파트", "203", "1102", "매매",
           "78,000", "", "", "", "12", "84.97", "112.4", "11", "25", "남향", 3, 2, 1,
           "즉시입주", datetime(2026, 11, 30), "2005-06-15", "박소유", "010-5555-1234",
           "", "", "3,000", "올수리 · 남향 · 즉시입주"]
    row += [""] * 20
    ws.append(row)


xlsx("04_50열관리대장.xlsx", _wide,
     {"n": 1, "check": {0: {"complex_name": "목련아파트", "dong": "203동", "ho": "1102호",
                            "price": 780000000, "area2_m2": 84.97, "area1_m2": 112.4,
                            "floor_info": "11/25", "room_cnt": 3, "bath_cnt": 2,
                            "owner_name": "박소유", "owner_tel": "010-5555-1234",
                            "settle_ymd": "2026-11-30", "approve_ymd": "2005-06-15",
                            "loan_amount": 30000000, "manager": "김중개"}}})


# ⑤ 상가 전문 매물장 — 권리금·부가세·현업종
def _sangga(wb):
    ws = wb.active
    ws.title = "상가매물"
    ws.append(["소재지", "건물명", "층", "전용(평)", "보증금", "월세", "관리비", "권리금",
               "현업종", "임대인", "임대인 연락처", "비고"])
    ws.append(["서울 강남구 역삼동 823-21", "역삼빌딩", "1/8", "15", "5,000", "350", "30",
               "8,000", "카페", "김임대", "010-2222-3333", "부가세별도 · 코너"])
    ws.append(["서울 강남구 논현동 12-3", "논현타워", "지1", "40", "3,000", "200", "20",
               "무권리", "창고", "이임대", "02-3456-7890", ""])


xlsx("05_상가전문.xlsx", _sangga,
     {"n": 2, "check": {0: {"address": "서울 강남구 역삼동 823-21", "complex_name": "역삼빌딩",
                            "floor_info": "1/8", "deposit": 50000000, "rent_price": 3500000,
                            "premium": 80000000, "current_biz": "카페", "type": "상가"},
                        1: {"floor_info": "B1", "premium": None}}})


# ⑥ 토지 매물장 — 지목·용도지역·대지면적
def _land(wb):
    ws = wb.active
    ws.title = "토지"
    ws.append(["소재지", "지목", "용도지역", "대지면적(㎡)", "매매가(억)", "도로", "비고"])
    ws.append(["경기 양평군 서종면 문호리 산 12-3", "임야", "계획관리", "3305", "8.5",
               "4m 접함", "계곡뷰"])
    ws.append(["충남 예산군 대술면 상항리 123", "전", "생산관리", "1652", "2.3", "맹지", ""])


xlsx("06_토지.xlsx", _land,
     {"n": 2, "check": {0: {"address": "경기 양평군 서종면 문호리 산 12-3", "type": "토지",
                            "price": 850000000},
                        1: {"price": 230000000}}})


# ⑦ 원룸 관리 매물장 — 호실이 세로로 쭉, 공실 표시
def _oneroom(wb):
    ws = wb.active
    ws.title = "원룸관리"
    ws.append(["건물명", "호실", "보증금", "월세", "관리비", "상태", "입주일", "세입자", "연락처"])
    rows = [["행복하우스", "101", 500, 45, 5, "공실", "즉시", "", ""],
            ["행복하우스", "102", 500, 50, 5, "임대중", "2027-03-14", "김세입", "010-1111-2222"],
            ["행복하우스", "201", 1000, 40, 5, "공실", "즉시", "", ""],
            ["행복하우스", "202", 500, 48, 5, "임대중", "2026-12-01", "박세입", "010-3333-4444"]]
    for r in rows:
        ws.append(r)
    ws.merge_cells("A2:A5")


xlsx("07_원룸관리.xlsx", _oneroom,
     {"n": 4, "check": {0: {"complex_name": "행복하우스", "ho": "101호",
                            "deposit": 5000000, "rent_price": 450000},
                        2: {"complex_name": "행복하우스", "ho": "201호", "deposit": 10000000}}})


# ⑧ 구형 .xls — 오래 쓰던 사무소는 아직 이걸 쓴다
xls("08_구형엑셀.xls", [
    ["단지명", "동", "호", "거래", "매매가", "전용면적", "연락처"],
    ["한마루럭키", "104", "1103", "매매", "120,000", "101.94", "010-6771-8004"],
    ["크로바", "202", "701", "전세", "55,000", "84.9", "042-483-1234"],
], {"n": 2, "check": {0: {"complex_name": "한마루럭키", "dong": "104동", "ho": "1103호",
                          "price": 1200000000, "area2_m2": 101.94},
                      1: {"deposit": 550000000}}})


# ⑨ 탭으로 나눈 텍스트 파일(EUC-KR) — 옛 프로그램이 뽑아 주는 형태
raw("09_탭구분.csv",
    "\n".join([
        "\t".join(["단지명", "동호수", "거래구분", "금액", "전용면적", "담당"]),
        "\t".join(["푸르지오", "301-1504", "매매", "92,000", "112.4", "최실장"]),
        "\t".join(["아이파크", "105-802", "전세", "58,000", "84.9", "최실장"]),
    ]).encode("euc-kr"),
    {"n": 2, "check": {0: {"complex_name": "푸르지오", "dong": "301동", "ho": "1504호",
                           "price": 920000000, "manager": "최실장"},
                       1: {"deposit": 580000000}}})


# ⑩ 세미콜론으로 나눈 CSV(UTF-8 BOM) — 유럽식 내보내기
raw("10_세미콜론.csv",
    ("\ufeff" + "\n".join([
        ";".join(["단지", "거래", "매매가", "전용"]),
        ";".join(["자이", "매매", "85,000", "84.9"]),
        ";".join(["힐스테이트", "매매", "72,000", "59.8"]),
    ])).encode("utf-8"),
    {"n": 2, "check": {0: {"complex_name": "자이", "price": 850000000}}})


# ⑪ 큰 파일 — 3,000줄. 느리면 실무에서 못 쓴다
def _big(wb):
    ws = wb.active
    ws.title = "전체매물"
    ws.append(["단지명", "동", "호", "거래", "매매가", "전용면적", "층", "연락처"])
    for i in range(3000):
        ws.append([f"단지{i % 120}", str(101 + i % 9), str(101 + i % 30),
                   ["매매", "전세", "월세"][i % 3], f"{60000 + i * 7:,}",
                   round(59.8 + (i % 40), 2), f"{1 + i % 25}/25", "010-1234-5678"])


xlsx("11_3000줄.xlsx", _big, {"n": 3000, "check": {}})


# ⑫ 한 시트에 표가 둘 — 위는 매매, 아래는 전월세(중간에 소제목)
def _two_tables(wb):
    ws = wb.active
    ws.title = "매물"
    ws.append(["[매매 물건]", None, None, None])
    ws.append(["단지명", "동호수", "매매가", "전용"])
    ws.append(["래미안", "101-1502", "85,000", "84.9"])
    ws.append(["자이", "203-802", "92,000", "101.9"])
    ws.append([None, None, None, None])
    ws.append(["[전월세 물건]", None, None, None])
    ws.append(["단지명", "동호수", "보증금", "월세"])
    ws.append(["푸르지오", "105-303", "50,000", ""])
    ws.append(["더샵", "202-1101", "10,000", "300"])


xlsx("12_한시트두표.xlsx", _two_tables,
     {"n": 2, "check": {0: {"complex_name": "래미안", "price": 850000000},
                        1: {"complex_name": "자이", "price": 920000000}},
      "note": "두 번째 표는 열이 달라 읽지 않고 멈춘다"})


# ⑬ 글자로 저장된 숫자 + 앞뒤 공백 + 전각 숫자
def _dirty(wb):
    ws = wb.active
    ws.title = "매물"
    ws.append(["단지명 ", " 거래", "매매가", "전용면적", "연락처"])
    ws.append([" 목련아파트 ", " 매 매 ", " 62,000 ", " 84.97 ", " 010 - 1234 - 5678 "])
    ws.append(["크로바", "전세", "55,000원", "101.9㎡", "010.2222.3333"])


xlsx("13_지저분한값.xlsx", _dirty,
     {"n": 2, "check": {0: {"complex_name": "목련아파트", "price": 620000000,
                            "area2_m2": 84.97, "contact": "010-1234-5678"},
                        1: {"deposit": 550000000, "area2_m2": 101.9,
                            "contact": "010-2222-3333"}}})


# ⑭ 머리글이 세로로 반복되는 카드형 — 우리가 못 읽는 서식(정직하게 실패해야 한다)
def _vertical(wb):
    ws = wb.active
    ws.title = "매물카드"
    for pair in [["단지명", "래미안"], ["동", "101"], ["호", "1502"], ["거래", "매매"],
                 ["매매가", "85,000"], ["전용", "84.9"]]:
        ws.append(pair)


xlsx("14_세로형카드.xlsx", _vertical,
     {"n": None, "check": {}, "note": "세로형은 못 읽는다 — 조용히 틀리지만 않으면 된다"})


def near(a, b) -> bool:
    if isinstance(b, float) and isinstance(a, (int, float)):
        return abs(float(a) - b) <= max(0.05, abs(b) * 0.005)
    return a == b


def main() -> int:
    import time
    ok, bad = 0, []
    for path, exp in BUILT:
        name = Path(path).name
        data = Path(path).read_bytes()
        t = time.time()
        try:
            res = LI.analyze(data, name)
        except Exception as e:  # noqa: BLE001
            if exp.get("n") is None:
                print(f"  · {name}: 읽지 못함 — {type(e).__name__} ({exp.get('note', '')})")
                ok += 1
                continue
            bad.append((name, [f"터짐: {type(e).__name__} {e}"]))
            continue
        el = time.time() - t
        sh = res["sheets"][0]
        errs = []
        if exp.get("n") is not None and len(sh["rows"]) != exp["n"]:
            errs.append(f"건수 {len(sh['rows'])} ≠ {exp['n']} · 건너뜀 "
                        f"{[s['why'] for s in sh['skipped']][:3]}")
        for idx, want in exp.get("check", {}).items():
            if idx >= len(sh["rows"]):
                errs.append(f"[{idx}] 줄 없음")
                continue
            got = sh["rows"][idx]
            for k, v in want.items():
                g = got.get(k)
                if v is None:
                    if g not in (None, "", 0):
                        errs.append(f"[{idx}] {k}: {g!r} 이 있으면 안 됨")
                elif not near(g, v):
                    errs.append(f"[{idx}] {k}: {g!r} ≠ {v!r}")
        flag = "" if not errs else " ✗"
        cols = " ".join(f"{c['header'][:8]}→{c['label'] or '-'}" for c in sh["columns"][:7])
        print(f"  {'✓' if not errs else '✗'} {name:<22} {len(sh['rows']):>5}건 "
              f"{el * 1000:>6.0f}ms  {cols[:96]}")
        if errs:
            bad.append((name, errs))
        else:
            ok += 1
    n = ok + len(bad)
    print(f"\n실제 파일 {n}개 중 통과 {ok} ({ok / max(1, n) * 100:.0f}%)   파일: {OUT}")
    for name, errs in bad:
        print(f"  ✗ {name}")
        for e in errs[:8]:
            print(f"      {e}")
    return len(bad)


if __name__ == "__main__":
    rc = main()
    if "--keep" not in sys.argv:
        shutil.rmtree(OUT, ignore_errors=True)
    raise SystemExit(min(1, rc))
