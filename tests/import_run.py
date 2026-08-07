#!/usr/bin/env python3
"""가져오기 채점기 — 코퍼스의 매물장을 진짜 엑셀로 구워 analyze() 에 넣고 기대값과 맞춘다.

    python3 tests/import_run.py            # xlsx
    python3 tests/import_run.py --xls      # 구형 .xls 로도(있으면)
    python3 tests/import_run.py --csv      # CSV(cp949)로도
"""
import io
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "scripts"))

import listing_import as LI  # noqa: E402
from tests.import_corpus import CASES  # noqa: E402


def to_xlsx(sheets, merge=()) -> bytes:
    import openpyxl
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, rows in sheets:
        ws = wb.create_sheet(title=name[:31])
        for r in rows:
            ws.append(list(r))
    for sheet_name, rng in merge:      # 병합 셀은 실제로 병합해 둬야 시험이 된다
        wb[sheet_name[:31]].merge_cells(rng)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_csv(sheets) -> bytes:
    import csv as _csv
    out = io.StringIO()
    w = _csv.writer(out)
    for r in sheets[0][1]:            # CSV 는 시트가 하나뿐이다
        w.writerow(["" if c is None else (c.strftime("%Y-%m-%d") if hasattr(c, "strftime") else c)
                    for c in r])
    return out.getvalue().encode("cp949", "replace")


def near(a, b) -> bool:
    if isinstance(b, float) and isinstance(a, (int, float)):
        return abs(float(a) - b) <= max(0.05, abs(b) * 0.005)
    return a == b


def run(fmt: str) -> int:
    ok_all, bad = 0, []
    for c in CASES:
        if fmt == "csv" and len(c["sheets"]) > 1:
            continue                   # CSV 는 시트가 하나라 다중 시트 사례는 건너뛴다
        if fmt == "csv" and c.get("merge"):
            continue                   # 병합 셀은 CSV 로 표현할 수 없다
        data = to_csv(c["sheets"]) if fmt == "csv" else to_xlsx(c["sheets"], c.get("merge") or ())
        fn = f"{c['name']}.{'csv' if fmt == 'csv' else 'xlsx'}"
        errs = []
        try:
            res = LI.analyze(data, fn)
            sh = res["sheets"][0]
        except Exception as e:  # noqa: BLE001
            bad.append((c["name"], [f"터짐: {type(e).__name__} {e}"]))
            continue
        exp = c["expect"]
        if len(sh["rows"]) != exp["n"]:
            errs.append(f"건수 {len(sh['rows'])} ≠ {exp['n']} "
                        f"(건너뜀 {sh['n_skipped']}: "
                        f"{[s['why'] for s in sh['skipped']][:4]})")
        if "skip" in exp and sh["n_skipped"] != exp["skip"]:
            errs.append(f"건너뜀 {sh['n_skipped']} ≠ {exp['skip']}")
        for idx, want in exp.get("check", {}).items():
            if idx >= len(sh["rows"]):
                errs.append(f"[{idx}] 줄이 없음")
                continue
            got = sh["rows"][idx]
            for k, v in want.items():
                g = got.get(k)
                if v is None:
                    if g not in (None, "", 0):
                        errs.append(f"[{idx}] {k}: {g!r} 이 있으면 안 됨")
                elif not near(g, v):
                    errs.append(f"[{idx}] {k}: {g!r} ≠ {v!r}")
        if errs:
            bad.append((c["name"], errs))
        else:
            ok_all += 1
    n = ok_all + len(bad)
    print(f"[{fmt}] {n}건 중 맞음 {ok_all} ({ok_all / max(1, n) * 100:.0f}%)")
    for name, errs in bad:
        print(f"  ✗ {name}")
        for e in errs[:6]:
            print(f"      {e}")
    return len(bad)


if __name__ == "__main__":
    fmts = ["xlsx"]
    if "--csv" in sys.argv:
        fmts = ["csv"]
    if "--all" in sys.argv:
        fmts = ["xlsx", "csv"]
    raise SystemExit(min(1, sum(run(f) for f in fmts)))
