#!/usr/bin/env python3
"""서울 59㎡ 12억 이하 500세대 이상 — 단지 정리 엑셀.

매물이 아니라 **단지**를 정리한다. 개인정보와 중개사 정보는 넣지 않는다 —
소유자·연락처·중개사무소·담당자는 원천에 있어도 이 파일에는 없다.

    python3 design/newsletter/make_seoul59_xlsx.py <full.json> <출력.xlsx>
"""
import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SRC, OUT = Path(sys.argv[1]), Path(sys.argv[2])
rows = json.loads(SRC.read_text(encoding="utf-8"))
ASOF = "2026-08-07"
PY = 3.3058

BRAND = "1268D3"
GROUPS = {  # 열 묶음별 머리 색 — 한 줄 머리글이라도 눈으로는 묶여 보이게
    "단지": "E8F0FF", "실거래": "FFF1E8", "호가": "EAF6EE", "활력": "F3EFFB", "위치": "EEF2F7",
}
# (머리글, 묶음, 값 함수, 서식, 폭)
def man(v):
    return round(v / 10000) if v else None


def pos(v):
    """0 은 '없음' 이 아니라 '모름' 인 칸들 — 건폐율 0%% 같은 값은 비워 둔다."""
    return v if v else None


COLS = [
    ("연번", "단지", lambda r, i: i, "0", 6),
    ("자치구", "단지", lambda r, i: r["gu"], None, 10),
    ("법정동", "단지", lambda r, i: r["dong"], None, 10),
    ("단지명", "단지", lambda r, i: r["nm"], None, 22),
    ("유형", "단지", lambda r, i: r["rtype"], None, 9),
    ("준공", "단지", lambda r, i: (f"{r['apr'][:4]}-{r['apr'][4:6]}" if r.get("apr") else None), None, 9),
    ("연차", "단지", lambda r, i: (2026 - int(r["apr"][:4])) if r.get("apr") else None, "0", 6),
    ("총세대수", "단지", lambda r, i: r["hh"], "#,##0", 9),
    ("동수", "단지", lambda r, i: pos(r["bld"]), "0", 6),
    ("최고층", "단지", lambda r, i: pos(r["hf"]), "0", 7),
    ("최저층", "단지", lambda r, i: pos(r["lf"]), "0", 7),
    ("용적률%", "단지", lambda r, i: pos(r["bcr"]), "0", 8),
    ("건폐율%", "단지", lambda r, i: pos(r["vlr"]), "0", 8),
    ("총주차", "단지", lambda r, i: pos(r["pk"]), "#,##0", 8),
    ("세대당주차", "단지", lambda r, i: pos(r["ppk"]), "0.00", 10),
    ("시공사", "단지", lambda r, i: r["builder"], None, 18),
    ("공급면적 최소㎡", "단지", lambda r, i: pos(r["mins"]), "0.0", 13),
    ("공급면적 최대㎡", "단지", lambda r, i: pos(r["maxs"]), "0.0", 13),

    ("전용㎡", "실거래", lambda r, i: r["ar"], "0.00", 8),
    ("전용평", "실거래", lambda r, i: round(r["ar"] / PY, 1) if r["ar"] else None, "0.0", 8),
    ("거래건수", "실거래", lambda r, i: r["n"], "0", 9),
    ("최저가(만원)", "실거래", lambda r, i: man(r["lo"]), "#,##0", 12),
    ("중앙값(만원)", "실거래", lambda r, i: man(r["med"]), "#,##0", 12),
    ("평균가(만원)", "실거래", lambda r, i: man(r["avg_amt"]), "#,##0", 12),
    ("최고가(만원)", "실거래", lambda r, i: man(r["hi"]), "#,##0", 12),
    ("전용평당가(만원)", "실거래",
     lambda r, i: round(man(r["med"]) / (r["ar"] / PY)) if r["ar"] and r["med"] else None, "#,##0", 14),
    ("첫거래일", "실거래", lambda r, i: r["first_ymd"], None, 11),
    ("마지막거래일", "실거래", lambda r, i: r["last_ymd"], None, 12),

    ("매매매물수", "호가", lambda r, i: r.get("an") or 0, "0", 10),
    ("매매최저호가(만원)", "호가", lambda r, i: man(r.get("alo")), "#,##0", 15),
    ("매매최고호가(만원)", "호가", lambda r, i: man(r.get("ahi")), "#,##0", 15),
    ("호가差%", "호가",
     lambda r, i: (round((r["alo"] / r["med"] - 1) * 100, 1) if r.get("alo") and r["med"] else None), "0.0", 9),
    ("전세매물수", "호가", lambda r, i: r.get("jn") or 0, "0", 10),
    ("전세최저호가(만원)", "호가", lambda r, i: man(r.get("jlo")), "#,##0", 15),
    ("갭(만원)", "호가",
     lambda r, i: (man(r["alo"]) - man(r["jlo"])) if r.get("alo") and r.get("jlo") else None, "#,##0", 11),

    ("단지 전체거래(6개월)", "활력", lambda r, i: r.get("n_all") or 0, "0", 15),
    ("세대수대비 거래율%", "활력",
     lambda r, i: round((r.get("n_all") or 0) / r["hh"] * 100, 1) if r["hh"] else None, "0.0", 15),
    ("단지 마지막거래일", "활력", lambda r, i: r.get("last_all"), None, 14),

    ("위도", "위치", lambda r, i: r["lat"], "0.000000", 11),
    ("경도", "위치", lambda r, i: r["lng"], "0.000000", 11),
    ("법정동코드", "위치", lambda r, i: r["cortar_no"], None, 12),
    ("단지코드", "위치", lambda r, i: r["cno"], None, 10),
    ("콕집 링크", "위치", lambda r, i: f"https://koczip.com/complex/{r['cno']}", None, 30),
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "단지목록"
thin = Side(style="thin", color="D8DEE7")
head_font = Font(name="맑은 고딕", size=9.5, bold=True, color="12203A")
body_font = Font(name="맑은 고딕", size=10)

ws.append([c[0] for c in COLS])
for j, (name, grp, _fn, _fmt, w) in enumerate(COLS, 1):
    cell = ws.cell(row=1, column=j)
    cell.font = head_font
    cell.fill = PatternFill("solid", fgColor=GROUPS[grp])
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(bottom=Side(style="medium", color=BRAND))
    ws.column_dimensions[get_column_letter(j)].width = w
ws.row_dimensions[1].height = 30

for i, r in enumerate(rows, 1):
    ws.append([fn(r, i) for _n, _g, fn, _f, _w in COLS])
    for j, (_n, _g, _fn, fmt, _w) in enumerate(COLS, 1):
        c = ws.cell(row=i + 1, column=j)
        c.font = body_font
        c.border = Border(bottom=thin)
        if fmt:
            c.number_format = fmt
            c.alignment = Alignment(horizontal="right")
ws.freeze_panes = "E2"          # 자치구·동·단지명까지 고정
ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{len(rows) + 1}"

# ── 안내 시트 — 이 숫자가 무엇인지 모르면 표는 위험하다 ────────────────────
g = wb.create_sheet("읽는 법")
g.column_dimensions["A"].width = 26
g.column_dimensions["B"].width = 96
guide = [
    ("무엇을 모았나", ""),
    ("대상", "서울 아파트 단지 중 ① 최근 6개월 전용 59㎡ 매매 실거래의 중앙값이 12억 이하이고 "
             "② 총 500세대 이상인 곳"),
    ("단지 수", f"{len(rows)}개 (자치구 {len({r['gu'] for r in rows})}곳, 합계 {sum(r['hh'] for r in rows):,}세대)"),
    ("기준일", ASOF),
    ("", ""),
    ("숫자를 어떻게 읽나", ""),
    ("금액 단위", "모두 만원입니다. 12억 = 120,000"),
    ("전용 59㎡", "단지마다 58.7·59.2·59.9㎡ 등으로 갈려 58.0~60.5㎡를 같은 타입으로 묶었습니다. "
                 "'전용㎡'는 그 단지 거래의 평균입니다"),
    ("중앙값", "가운데 값입니다. 한두 건의 특이 거래에 휘둘리지 않게 평균 대신 이걸 기준으로 삼았습니다. "
              "평균가는 참고로 같이 실었습니다"),
    ("거래건수", "3건 이하면 중앙값도 흔들립니다 — 참고치로만 보십시오"),
    ("전용평당가", "중앙값 ÷ (전용㎡ ÷ 3.3058). 공급면적 기준이 아니라 전용 기준입니다"),
    ("호가差%", "현재 가장 싼 매물 호가가 실거래 중앙값보다 얼마나 위인지. 클수록 지금 사려면 더 얹어야 합니다"),
    ("갭", "매매 최저호가 − 전세 최저호가. 둘 다 호가라 실제 필요한 돈과 다를 수 있습니다"),
    ("세대수대비 거래율", "최근 6개월 그 단지 전 면적 거래건수 ÷ 총세대수. 단지가 얼마나 도는지 봅니다"),
    ("", ""),
    ("주의", ""),
    ("신고 지연", "실거래는 계약 후 최대 30일까지 신고할 수 있어 최근 한 달치는 아직 다 들어오지 않았습니다"),
    ("해제거래", "해제(취소)된 거래는 뺐습니다. 넣으면 없던 신고가가 시세처럼 섞입니다"),
    ("호가", "지금 광고 중인 매물 기준이라 실제 거래 가능 가격과 다를 수 있습니다. 0이면 매물이 없다는 뜻입니다"),
    ("빈 칸", "원천에 값이 없는 항목입니다. 짐작해서 채우지 않았습니다"),
    ("난방방식", "원천이 코드(HT001 등)로만 주고 무엇을 뜻하는지 확인되지 않아 넣지 않았습니다"),
    ("", ""),
    ("개인정보", ""),
    ("넣지 않은 것", "소유자·임대인·세입자 이름과 연락처, 중개사무소 이름·등록번호·연락처, 담당자, "
                   "매물 단위 정보. 이 파일은 단지 정보와 집계값만 담습니다"),
    ("", ""),
    ("출처", "국토교통부 아파트 매매 실거래가 · 단지 정보 및 매물 호가 · 집계 콕집(koczip.com)"),
]
for i, (k, v) in enumerate(guide, 1):
    g.cell(row=i, column=1, value=k).font = Font(name="맑은 고딕", size=10, bold=True,
                                                 color=BRAND if not v else "12203A")
    c = g.cell(row=i, column=2, value=v)
    c.font = Font(name="맑은 고딕", size=10)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    g.row_dimensions[i].height = 30 if len(v) > 60 else 16

wb.save(OUT)
print(f"{OUT} · 단지 {len(rows)} · 열 {len(COLS)}")
